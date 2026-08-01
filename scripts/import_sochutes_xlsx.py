# -*- coding: utf-8 -*-
"""Importa a base ALL-TIME do Só Chutes (.xlsx, aba 'Base') para dono='SoChutes'.

Conta criada na s223 (`app/auth.py`, dono SOLO) para o tipster Só Chutes — o
primeiro tipster servido pela plataforma. Esta base é o histórico da planilha
dele (17/09/2024 → 27/07/2026, 23.199 apostas); o planilhamento novo passa a
ser feito pelo bot Sharpen daqui em diante.

Layout da planilha (header na linha 7, dados = linhas com data real na col. 0):

    0 DATA | 1 MERCADO | 2 JOGADOR | 3 TIPO | 4 STAKE | 5 ODD | 6 RESULTADO |
    7 PROFIT | 8 MÊS | 9 ANO | 10 SEMANA | 11 BOOKIE

Decisões do Feca (sessão 223):

- **Casa**: `BOOKIE` só existe a partir de 01/04/2026 (Bet365/Superbet/Betano).
  Linha sem BOOKIE entra como **Bet365** (padrão do tipster).
- **Stake em UNIDADES**: a planilha conta em u (0,05–3,25). Importa 1u = 1 —
  o P/L do dashboard é o P/L em unidades (a base fecha em +1.382,04u).
- **Odd inválida com resultado L → odd 1**: 56 linhas têm odd 0 (todas L).
  P/L de L não usa a odd (−stake), e odd 1 mantém o estado 'resolvida' em vez
  de pendurar pendência eterna em Abertas.
- **Odd que o Excel comeu**: célula que virou data é a odd digitada com ponto
  no locale BR ("6.5" → 06/05/2025); reconstrói `dia.mês`. "18*72" → 18,72.
  Todas essas linhas são L/V — a recuperação é cosmética, o P/L não muda.
- **Dono solo**: sem OPERADORES, sem dedup cruzada.

Validação prévia (inspect_sochutes2.py): as 3.142 vitórias fecham
`profit = stake × (odd − 1)` sem exceção — as odds >100 são triplas/trixies
reais, não erro de digitação. L fecha `profit = −stake` e V fecha `profit = 0`
em 100 % das linhas.

Uso:
    python scripts/import_sochutes_xlsx.py --xlsx "C:\\...\\Só Chutes - All-time.xlsx"        # DRY
    python scripts/import_sochutes_xlsx.py --xlsx "C:\\...\\Só Chutes - All-time.xlsx" --go   # escreve
"""
import argparse
import asyncio
import datetime as dt
import hashlib
import os
import re
from collections import Counter

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
DONO = 'SoChutes'
PARCEIRO = 'Padrão'
TIPSTER = 'Só Chutes'
ESPORTE = 'Futebol'          # base 100 % futebol (mercados de jogador)
CASA_DEFAULT = 'Bet365'      # BOOKIE vazia = Bet365 (decisão do Feca)
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}


# ---------- sanitização de texto ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


# ---------- casa ----------
# `casa` é TEXTO em 7 tabelas: cada grafia é uma casa DIFERENTE (CLAUDE.md).
# As três grafias da planilha já são as canônicas do sistema.
_CASA_MAP = {'BET365': 'Bet365', 'SUPERBET': 'Superbet', 'BETANO': 'Betano'}


def norm_casa(v) -> str:
    bruto = limpa(v)
    if not bruto:
        return CASA_DEFAULT
    return _CASA_MAP.get(bruto.upper().replace(' ', ''), bruto)


# ---------- categoria + descrição ----------
# Combo (Dupla/Tripla/Trixie) → `Múltipla` (MASTER_APOSTAS §5: cupom com várias
# seleções é Múltipla mesmo quando todas são do mesmo mercado). O mercado real
# vai preservado como sufixo da descrição — categoria Múltipla sozinha perderia
# a informação de que o cupom é de Anytime.
_TIPOS_COMBO = {'dupla / tripla / trixie', 'dupla', 'multipla'}

# mercado → (categoria quando SIMPLES, sufixo da descrição quando simples,
#            sufixo da descrição quando COMBO)
# HT - Anytime é mercado de MOMENTO (marcar no 1º tempo): a categoria segue
# `Anytime` (MASTER_APOSTAS — limiar/momento não muda categoria) e o momento
# vai na descrição, senão o mesmo jogador em HT e FT colapsa em texto igual.
_MERCADOS = {
    'HT - Anytime':               ('Anytime',      ' - 1º Tempo', ' - Anytime 1º Tempo'),
    'FT - Anytime':               ('Anytime',      '',            ' - Anytime'),
    'Assistência':                ('Assistência',  '',            ' - Assistência'),
    'Dupla/Tripla Any':           ('Anytime',      '',            ' - Anytime'),
    'FT - Anytime / Assistência': ('Anytime',      '',            ' - Anytime + Assistência'),
    'Finalizações':               ('Chutes',       '',            ' - Finalizações'),
    'Faltas':                     ('Player Props', ' - Faltas',   ' - Faltas'),
    'Outros':                     ('Outros',       '',            ''),
}


def eh_combo(tipo, jogador: str) -> bool:
    """Cupom com mais de uma seleção. O TIPO manda; '/' no jogador pega as 34
    linhas antigas rotuladas 'Simples' com vários jogadores na célula."""
    return limpa(tipo).lower() in _TIPOS_COMBO or '/' in jogador


def categoria(mercado, combo: bool) -> str:
    if combo:
        return 'Múltipla'
    cat, _, _ = _MERCADOS.get(limpa(mercado), ('Outros', '', ''))
    return cat


def descricao(mercado, jogador: str, combo: bool) -> str:
    """Jogador(es) + sufixo do mercado. ' // ' é o ÚNICO separador de seleção
    (regra #19) — a barra da planilha vira ' // '."""
    d = jogador
    if combo:
        d = re.sub(r'\s*/\s*', ' // ', d)
    _, suf_simples, suf_combo = _MERCADOS.get(limpa(mercado), ('Outros', '', ''))
    return d + (suf_combo if combo else suf_simples)


# ---------- data / stake / odd ----------
def norm_data(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def _para_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = limpa(v).replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def fmt_stake(v) -> str:
    n = _para_float(v)
    return '' if n is None else f'{n:.2f}'.replace('.', ',')


def _float_str(n: float) -> str:
    """Precisão completa (nunca truncar odd), vírgula decimal."""
    s = repr(float(n))
    if s.endswith('.0'):
        s = s[:-2]
    return s.replace(',', '').replace('.', ',')


def norm_odd(v, resultado: str) -> str:
    """Odd da planilha com as recuperações decididas na s223 (ver docstring)."""
    if isinstance(v, dt.datetime):
        n = float(f'{v.day}.{v.month}')          # "6.5" digitado → 06/05/2025
    elif isinstance(v, str) and '*' in v:
        n = _para_float(v.replace('*', '.'))     # "18*72" → 18.72
    else:
        n = _para_float(v)
    if n is None or n <= 0:
        return '1' if resultado == 'L' else ''   # decisão do Feca: L sem odd → 1
    return _float_str(n)


def norm_resultado(v) -> str:
    r = limpa(v).upper()
    return r if r in VALID else ''


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if 'Base' not in wb.sheetnames:
        raise SystemExit(f"aba 'Base' não encontrada — abas: {wb.sheetnames}")
    ws = wb['Base']
    # linha de dado = data REAL na col. 0 (descarta o bloco de KPIs do topo,
    # o cabeçalho e as linhas vazias do export)
    brutas = [r for r in ws.iter_rows(values_only=True)
              if r and isinstance(r[0], dt.datetime)]
    out = []
    for r in brutas:
        jogador = limpa(r[2])
        combo = eh_combo(r[3], jogador)
        resultado = norm_resultado(r[6])
        out.append({
            'data': norm_data(r[0]),
            'esporte': ESPORTE,
            'tipster': TIPSTER,
            'casa': norm_casa(r[11]),
            'parceiro': PARCEIRO,
            'aposta': categoria(r[1], combo),
            'descricao': descricao(r[1], jogador, combo),
            'stake': fmt_stake(r[4]),
            'odd': norm_odd(r[5], resultado),
            'resultado': resultado,
        })
    return out


# ---------- assinatura (idêntica a repository._assinatura / _norm_odd) ----------
# ATENÇÃO: `stake` ENTRA no hash. Não copiar `import_lava.py` /
# `import_dashboard_xlsx.py` (versão pré-s133, sem stake). Divergir daqui faz a
# próxima captura gerar assinatura nova e duplicar o histórico inteiro.
def _norm_odd_sig(v: str) -> str:
    try:
        return f"{round(float(v.replace(',', '.')), 2):.2f}"
    except (ValueError, AttributeError):
        return v


def assinaturas(rows: list[dict]) -> list[str]:
    """Assinatura por linha com contador de duplicatas por lote (igual ao app):
    conteúdo idêntico ESCALA o contador, nunca colide."""
    counts: dict[str, int] = {}
    sigs = []
    for r in rows:
        base_raw = "|".join([
            r['casa'], r['parceiro'], r['data'], r['aposta'], r['descricao'],
            r['stake'], _norm_odd_sig(r['odd']),
        ])
        base_sig = hashlib.sha256(base_raw.encode()).hexdigest()[:20]
        cnt = counts.get(base_sig, 0) + 1
        counts[base_sig] = cnt
        raw = base_raw if cnt == 1 else f"{base_raw}|{cnt}"
        sigs.append(hashlib.sha256(raw.encode()).hexdigest()[:20])
    return sigs


def estado_extracao(resultado: str, odd: str) -> str:
    """Espelha repository.estado_extracao: 'resolvida' EXIGE resultado canônico
    E odd > 0."""
    if resultado not in VALID:
        return 'aberta'
    n = _para_float(odd)
    return 'resolvida' if (n or 0) > 0 else 'aberta'


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict]):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    sigs = assinaturas(rows)
    casas = sorted({r['casa'] for r in rows})

    registros = []
    for r, sig in zip(rows, sigs):
        registros.append((
            DONO, r['casa'], r['parceiro'], sig, None,           # codigo_bilhete
            r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
            r['stake'], r['odd'], r['resultado'] or None,
            estado_extracao(r['resultado'], r['odd']),
            None, None, ORIGEM,                                  # confianca, stake_usd
        ))

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=120)
            try:
                async with conn.transaction():
                    # idempotente: reimportar não acumula (limpa só o que ESTE
                    # import escreveu — captura do bot/extensão tem outra origem)
                    apagadas = await conn.execute(
                        "DELETE FROM bilhetes WHERE dono=$1 AND origem=$2", DONO, ORIGEM)
                    print(f'  [tentativa {tentativa}] limpou import anterior: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    for casa in casas:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            DONO, casa, PARCEIRO)
                    # feed ordena por criado_em DESC; num import não existe
                    # "envio" → ancora na data da aposta para sair cronológico
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC,
                                                               id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem=$2
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, DONO, ORIGEM)
                n = await conn.fetchval("SELECT COUNT(*) FROM bilhetes WHERE dono=$1", DONO)
                nc = await conn.fetchval(
                    "SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1", DONO)
                np = await conn.fetchval("SELECT COUNT(*) FROM parceiros WHERE dono=$1", DONO)
                print(f'\nOK — bilhetes dono={DONO}={n} | casas={nc} | contas={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


# ---------- relatório do DRY ----------
def _pl_derivado(stake: str, odd: str, res: str):
    s = _para_float(stake) or 0.0
    o = _para_float(odd)
    if res == 'L':
        return -s
    if res == 'V':
        return 0.0
    if o is None or o <= 0:
        return None
    if res == 'W':
        return s * (o - 1)
    if res == 'HW':
        return s * (o - 1) / 2
    if res == 'HL':
        return -s / 2
    return None


def _relatorio(rows: list[dict]):
    print(f'DONO={DONO} | conta={PARCEIRO!r} | linhas: {len(rows)}')
    for campo in ('casa', 'esporte', 'tipster', 'aposta'):
        print(f'\n{campo}:', dict(Counter(r[campo] for r in rows).most_common()))
    print('\nresultado:',
          dict(Counter(r['resultado'] or '(aberta)' for r in rows).most_common()))

    sigs = assinaturas(rows)
    dup = len(sigs) - len(set(sigs))
    print(f'\nassinaturas: {len(set(sigs))} únicas de {len(sigs)} '
          f'({dup} desambiguadas por contador)')

    estados = Counter(estado_extracao(r['resultado'], r['odd']) for r in rows)
    print('extraction_state:', dict(estados))

    sem_odd = [r for r in rows if not r['odd']]
    if sem_odd:
        print(f'⚠ {len(sem_odd)} linha(s) sem odd (ficam abertas):')
        for r in sem_odd[:10]:
            print(f'    {r["data"]} | {r["resultado"] or "—"} | {r["descricao"][:52]}')

    sem_stake = [r for r in rows if (_para_float(r['stake']) or 0) <= 0]
    if sem_stake:
        print(f'\n⚠ {len(sem_stake)} linha(s) com stake 0: gravadas, porém INVISÍVEIS '
              f'no dashboard (dashboard_rows corta stake <= 0).')
        for r in sem_stake:
            print(f'    {r["data"]} | {r["casa"]} | {r["descricao"][:52]}')

    turnover = sum(_para_float(r['stake']) or 0 for r in rows)
    pl = sum(v for r in rows
             if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is not None)
    print(f'\nturnover (u):     {turnover:>14,.2f}   (planilha: 12.182,20)')
    print(f'P/L derivado (u): {pl:>14,.2f}   (planilha: 1.382,04)')

    print('\n=== 16 AMOSTRAS (Data|Esporte|Tipster|Casa|Conta|Aposta|Descrição|Stake|Odd|Res) ===')
    step = max(1, len(rows) // 16)
    for r in rows[::step][:16]:
        print(' | '.join([r['data'], r['esporte'], r['tipster'], r['casa'], r['parceiro'],
                          r['aposta'], r['descricao'][:44], r['stake'], r['odd'],
                          r['resultado'] or '—']))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows))


if __name__ == '__main__':
    main()
