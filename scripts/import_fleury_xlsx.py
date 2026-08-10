# -*- coding: utf-8 -*-
"""Importa a base do tipster Fleury (FleuryChutes.xlsx, aba 'Geral') para dono='Fleury'.

Base de nicho: 100 % mercados de FINALIZAÇÃO no futebol (11/06/2026 → 09/08/2026,
473 apostas). Conta criada nesta sessão em `app/auth.py` (dono SOLO).

Layout da planilha (header na linha 1; as colunas J..V são painéis de KPI da
própria planilha e são ignoradas):

    0 Data | 1 Casa | 2 Jogador/Equipe | 3 Mercado | 4 Odd | 5 Stake |
    6 Resultado | 7 Lucro

Por que a aba 'Geral' e não as três mensais (06.26 / 07.26 / 08.26):

    'Geral' é a união EXATA das três (178 + 207 + 88 = 473) e, em 3 linhas de
    15/06, traz o nome completo que as mensais cortam ('Ghoddos' → 'Saman
    Ghoddos'). Conferido linha a linha por multiconjunto: 473 = 473, e a única
    diferença são essas 3 grafias.

Decisões do Feca (esta sessão):

- **Dono novo `Fleury`**, solo — sem OPERADORES, sem dedup cruzada. O login só
  passa a funcionar quando `SENHA_FLEURY_HASH` for colada no Railway.
- **Tipster** = `Fleury` em todas as linhas.
- **Uma conta por casa**: `Padrão` em cada uma das 4 casas (Bet365, Betano,
  Superbet, BetMGM) → 4 linhas no Painel de Contas, cada uma com custo próprio.
- **Stake em UNIDADES**: a planilha conta em u (0,10–4,00). Importa 1u = 1 — o
  P/L do dashboard é o P/L em unidades (a base fecha em +122,30u sobre 447,80u
  de turnover).

Classificação (o `Mercado` da planilha diz 'Chutes' nas 473 linhas; o MASTER
distingue três casos, e é ele que manda):

- **`Múltipla`** para cupom com mais de uma seleção (`MASTER_APOSTAS §Bet
  Builder`: Bet Builder é Múltipla *mesmo quando todas as seleções são do mesmo
  jogo*). Detectado por prefixo `Dupla`/`Tripla`, `/`, `,` ou ` + ` unindo
  seleções. As seleções são separadas por ` // ` — o ÚNICO separador (regra #19).
- **`Chutes no Gol`** para finalização NO ALVO (`MASTER_APOSTAS §9.14`: nunca
  usar para total de finalizações). São as marcas `SoG`, `OBSoG` e `no gol`.
- **`Chutes`** para o resto — total de finalizações, incluindo os recortes que
  NÃO mudam a categoria (`cabeça`, `HT`, handicap de time como `Bósnia o8.5`).

Cuidado ao mexer no detector de categoria: `cab` casa com sobrenome
(`Bruno Cabrera`, `Jovane Cabral`, `Jovane Cabral`) — por isso `cabeça`/`cab.`
são casados com fronteira, e `cabeça` é `Chutes`, não `Chutes no Gol`, exceto
quando vem com `no gol`.

Normalizações de resultado: `VOID` → `V` (3 linhas), `l` minúsculo → `L`
(2 linhas). As 473 linhas fecham `lucro = stake × (odd − 1) | −stake | 0` sem
uma única divergência — a base é consistente e não precisa de recuperação de
odd (nenhuma linha sem odd, nenhuma sem resultado).

Uso:
    python scripts/import_fleury_xlsx.py --xlsx "C:\\Users\\...\\FleuryChutes.xlsx"        # DRY
    python scripts/import_fleury_xlsx.py --xlsx "C:\\Users\\...\\FleuryChutes.xlsx" --go   # escreve
"""
import argparse
import asyncio
import datetime as dt
import hashlib
import os
import re
from collections import Counter

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
DONO = 'Fleury'
PARCEIRO = 'Padrão'
TIPSTER = 'Fleury'
ESPORTE = 'Futebol'          # base 100 % futebol (finalizações de jogador/time)
ABA = 'Geral'
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}


# ---------- sanitização de texto ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


# ---------- casa ----------
# `casa` é TEXTO em 7 tabelas: cada grafia é uma casa DIFERENTE (CLAUDE.md).
# As 4 grafias da planilha já são as canônicas do sistema (as mesmas do
# `_CASA_DISPLAY` / dos mapas de favicon). Casa desconhecida entra VERBATIM —
# nunca title-casear, que mutila nome e cria conta paralela.
_CASA_MAP = {
    'BET365': 'Bet365',
    'BETANO': 'Betano',
    'SUPERBET': 'Superbet',
    'BETMGM': 'BetMGM',
}


def norm_casa(v) -> str:
    bruto = limpa(v)
    return _CASA_MAP.get(bruto.upper().replace(' ', ''), bruto)


# ---------- categoria + descrição ----------
# Cupom com mais de uma seleção. Quatro marcas, e nenhuma delas é o '+' de
# limiar ('2+' = 2 ou mais finalizações) — só o ' + ' cercado de espaço une
# seleções ('Jashari FdA + Zakaria cabeça').
_PREFIXO_COMBO = re.compile(r'^(dupla|tripla|trixie)\b', re.I)
_SEPARADORES = re.compile(r'\s*/\s*|\s*,\s*|\s+\+\s+')

# Finalização NO ALVO. 'cab. no gol' entra; 'cabeça' sozinho NÃO (é chute de
# cabeça, total). Fronteira à esquerda para não casar sobrenome.
_NO_ALVO = re.compile(r'\bO?BSoG\b|\bSoG\b|no gol|no alvo', re.I)


def eh_combo(desc: str) -> bool:
    return bool(_PREFIXO_COMBO.search(desc) or _SEPARADORES.search(desc))


def categoria(desc: str) -> str:
    if eh_combo(desc):
        return 'Múltipla'
    return 'Chutes no Gol' if _NO_ALVO.search(desc) else 'Chutes'


def descricao(desc: str) -> str:
    """Texto verbatim da planilha, com as seleções separadas por ' // ' —
    o ÚNICO separador de seleção (regra #19), inclusive em mesmo-jogo."""
    if not eh_combo(desc):
        return desc
    partes = [p.strip() for p in _SEPARADORES.split(desc) if p.strip()]
    return ' // '.join(partes) if len(partes) > 1 else desc


# ---------- data / stake / odd ----------
def norm_data(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def _para_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = limpa(v).replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def fmt_stake(v) -> str:
    n = _para_float(v)
    return '' if n is None else f'{n:.2f}'.replace('.', ',')


def _float_str(n: float) -> str:
    """Precisão completa (nunca truncar odd), vírgula decimal."""
    s = repr(float(n))
    if s.endswith('.0'):
        s = s[:-2]
    return s.replace(',', '').replace('.', ',')


def norm_odd(v) -> str:
    n = _para_float(v)
    return '' if n is None or n <= 0 else _float_str(n)


# `VOID` é a grafia da planilha para anulada; `l` minúsculo é digitação.
# O `.upper()` sozinho já resolve os dois — o mapa deixa a intenção explícita.
_RESULTADO_MAP = {'VOID': 'V', 'ANULADA': 'V', 'CANCELADA': 'V'}


def norm_resultado(v) -> str:
    r = limpa(v).upper()
    r = _RESULTADO_MAP.get(r, r)
    return r if r in VALID else ''


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        raise SystemExit(f"aba {ABA!r} não encontrada — abas: {wb.sheetnames}")
    ws = wb[ABA]
    # linha de dado = data REAL na col. 0 (descarta cabeçalho, linhas vazias e
    # os painéis de KPI que a planilha mantém nas colunas J..V)
    brutas = [r for r in ws.iter_rows(values_only=True)
              if r and isinstance(r[0], dt.datetime)]
    out = []
    for r in brutas:
        bruto = limpa(r[2])
        out.append({
            'data': norm_data(r[0]),
            'esporte': ESPORTE,
            'tipster': TIPSTER,
            'casa': norm_casa(r[1]),
            'parceiro': PARCEIRO,
            'aposta': categoria(bruto),
            'descricao': descricao(bruto),
            'stake': fmt_stake(r[5]),
            'odd': norm_odd(r[4]),
            'resultado': norm_resultado(r[6]),
            '_lucro': _para_float(r[7]),      # só para a conferência do DRY
            '_bruto': bruto,
        })
    return out


# ---------- assinatura (idêntica a repository._assinatura / _norm_odd) ----------
# ATENÇÃO: `stake` ENTRA no hash. Não copiar `import_lava.py` /
# `import_dashboard_xlsx.py` (versão pré-s133, sem stake). Divergir daqui faz a
# próxima captura gerar assinatura nova e duplicar o histórico inteiro.
def _norm_odd_sig(v: str) -> str:
    try:
        return f"{round(float(v.replace(',', '.')), 2):.2f}"
    except (ValueError, AttributeError):
        return v


def assinaturas(rows: list[dict]) -> list[str]:
    """Assinatura por linha com contador de duplicatas por lote (igual ao app):
    conteúdo idêntico ESCALA o contador, nunca colide."""
    counts: dict[str, int] = {}
    sigs = []
    for r in rows:
        base_raw = "|".join([
            r['casa'], r['parceiro'], r['data'], r['aposta'], r['descricao'],
            r['stake'], _norm_odd_sig(r['odd']),
        ])
        base_sig = hashlib.sha256(base_raw.encode()).hexdigest()[:20]
        cnt = counts.get(base_sig, 0) + 1
        counts[base_sig] = cnt
        raw = base_raw if cnt == 1 else f"{base_raw}|{cnt}"
        sigs.append(hashlib.sha256(raw.encode()).hexdigest()[:20])
    return sigs


def estado_extracao(resultado: str, odd: str) -> str:
    """Espelha repository.estado_extracao: 'resolvida' EXIGE resultado canônico
    E odd > 0."""
    if resultado not in VALID:
        return 'aberta'
    n = _para_float(odd)
    return 'resolvida' if (n or 0) > 0 else 'aberta'


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict]):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    sigs = assinaturas(rows)
    casas = sorted({r['casa'] for r in rows})

    registros = []
    for r, sig in zip(rows, sigs):
        registros.append((
            DONO, r['casa'], r['parceiro'], sig, None,           # codigo_bilhete
            r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
            r['stake'], r['odd'], r['resultado'] or None,
            estado_extracao(r['resultado'], r['odd']),
            None, None, ORIGEM,                                  # confianca, stake_usd
        ))

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=120)
            try:
                async with conn.transaction():
                    # idempotente: reimportar não acumula (limpa só o que ESTE
                    # import escreveu — captura do bot/extensão tem outra origem)
                    apagadas = await conn.execute(
                        "DELETE FROM bilhetes WHERE dono=$1 AND origem=$2", DONO, ORIGEM)
                    print(f'  [tentativa {tentativa}] limpou import anterior: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    # uma conta `Padrão` POR CASA (decisão do Feca): 4 linhas no
                    # Painel de Contas, cada uma com custo próprio
                    for casa in casas:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            DONO, casa, PARCEIRO)
                    # feed ordena por criado_em DESC; num import não existe
                    # "envio" → ancora na data da aposta para sair cronológico
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC,
                                                               id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem=$2
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, DONO, ORIGEM)
                n = await conn.fetchval("SELECT COUNT(*) FROM bilhetes WHERE dono=$1", DONO)
                nc = await conn.fetchval(
                    "SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1", DONO)
                np = await conn.fetchval("SELECT COUNT(*) FROM parceiros WHERE dono=$1", DONO)
                print(f'\nOK — bilhetes dono={DONO}={n} | casas={nc} | contas={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


# ---------- relatório do DRY ----------
def _pl_derivado(stake: str, odd: str, res: str):
    s = _para_float(stake) or 0.0
    o = _para_float(odd)
    if res == 'L':
        return -s
    if res == 'V':
        return 0.0
    if res == 'HL':
        return -s / 2
    if o is None:
        return None
    if res == 'W':
        return s * (o - 1)
    if res == 'HW':
        return s * (o - 1) / 2
    return None


def _relatorio(rows: list[dict]):
    print(f'DONO={DONO} | tipster={TIPSTER!r} | conta={PARCEIRO!r} por casa | '
          f'linhas: {len(rows)}')
    for campo in ('casa', 'esporte', 'tipster', 'aposta'):
        print(f'\n{campo}:', dict(Counter(r[campo] for r in rows).most_common()))
    print('\nresultado:',
          dict(Counter(r['resultado'] or '(aberta)' for r in rows).most_common()))

    datas = sorted(dt.datetime.strptime(r['data'], '%d/%m/%Y') for r in rows)
    print(f"período: {datas[0]:%d/%m/%Y} → {datas[-1]:%d/%m/%Y}")

    sigs = assinaturas(rows)
    dup = len(sigs) - len(set(sigs))
    print(f'\nassinaturas: {len(set(sigs))} únicas de {len(sigs)} '
          f'({dup} desambiguadas por contador)')

    estados = Counter(estado_extracao(r['resultado'], r['odd']) for r in rows)
    print('extraction_state:', dict(estados))

    sem_odd = [r for r in rows if not r['odd']]
    if sem_odd:
        print(f'⚠ {len(sem_odd)} linha(s) sem odd (ficam abertas):')
        for r in sem_odd[:10]:
            print(f'    {r["data"]} | {r["resultado"] or "—"} | {r["descricao"][:52]}')

    sem_stake = [r for r in rows if (_para_float(r['stake']) or 0) <= 0]
    if sem_stake:
        print(f'\n⚠ {len(sem_stake)} linha(s) com stake 0: gravadas, porém INVISÍVEIS '
              f'no dashboard (dashboard_rows corta stake <= 0).')

    # conferência contra a coluna Lucro da própria planilha, linha a linha:
    # se o P/L derivado divergir, a classificação ou a normalização quebrou
    div = [r for r in rows
           if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is None
           or abs(v - (r['_lucro'] or 0)) > 0.02]
    print(f'\nP/L derivado × coluna Lucro da planilha: {len(div)} divergência(s)')
    for r in div[:10]:
        print(f'    {r["data"]} | {r["resultado"]} | {r["descricao"][:44]} | '
              f'planilha={r["_lucro"]}')

    turnover = sum(_para_float(r['stake']) or 0 for r in rows)
    pl = sum(v for r in rows
             if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is not None)
    print(f'\nturnover (u):     {turnover:>10,.2f}   (planilha: 447,80)')
    print(f'P/L derivado (u): {pl:>10,.2f}   (planilha: 122,30)')

    recat = [r for r in rows if r['aposta'] != 'Chutes']
    print(f'\n=== {len(recat)} LINHA(S) FORA DE `Chutes` (o MASTER manda, não a '
          f'coluna Mercado) ===')
    for r in recat:
        print(f'    [{r["aposta"]:>13}] {r["_bruto"]}'
              + ('' if r['descricao'] == r['_bruto'] else f'   →   {r["descricao"]}'))

    print('\n=== 16 AMOSTRAS (Data|Esporte|Tipster|Casa|Conta|Aposta|Descrição|Stake|Odd|Res) ===')
    step = max(1, len(rows) // 16)
    for r in rows[::step][:16]:
        print(' | '.join([r['data'], r['esporte'], r['tipster'], r['casa'], r['parceiro'],
                          r['aposta'], r['descricao'][:44], r['stake'], r['odd'],
                          r['resultado'] or '—']))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows))


if __name__ == '__main__':
    main()
