# -*- coding: utf-8 -*-
"""Importa a base do arrudex (`2026 - Fernando.xlsx`, aba 'Apostas') para dono='arrudex'.

⚠️ O `dono` é o USERNAME, conferido na tabela `usuarios` antes do import — não
deduzido do nome do arquivo (que é o template da planilha, não a pessoa): username
`arrudex`, status `ativo`, hash de 60 caracteres, e-mail `raphaelarruda96@gmail.com`,
criado em 17/08/2026. Cadastro em autosserviço (Fase 2), aprovado pelo Feca: **não
existe env var nem linha em `app/auth.py` para esta conta**. A base estava vazia
(0 bilhetes) antes deste import.

Que a planilha NÃO é a do Feca foi MEDIDO, não suposto: o overlap de (data,
descrição) com os 30.979 bilhetes do dono `Feca` deu **0 de 7.110 chaves**. As
contas também não batem (lá são e-mails com operador entre colchetes; aqui são
nomes próprios), embora os tipsters seguidos sejam quase os mesmos — é o mesmo
círculo, operação diferente.

── Layout da planilha ────────────────────────────────────────────────────────

Header na linha 1; linha de dado = `Data` REAL na coluna 0 (mesmo critério do
`import_fleury_xlsx.py` / `import_passatips_xlsx.py`). **Só as colunas A..L são
apostas** — da M em diante há painéis de KPI da própria planilha (listas de
Resultado/Casa/Conta/Esporte e placares por tipster) que são IGNORADOS; lê-los
criaria linha fantasma:

    0 Data | 1 Tipster | 2 Aposta (descrição) | 3 Odd | 4 Stake (R$) | 5 u |
    6 Resultado | 7 PL R$ | 8 PL u | 9 Casa | 10 Conta | 11 Esporte

10.981 apostas, 01/01/2026 → 22/08/2026 (as datas além de hoje são apostas em
aberto de evento futuro). 34 tipsters, 37 casas, 36 contas, 11 rótulos de esporte.
Nenhuma linha sem casa, conta, tipster ou stake.

── Decisões do Feca (esta sessão) ────────────────────────────────────────────

- **Base inteira** (10.981), todos os tipsters — é a carteira dele, e `Arrudex`
  aparece lá dentro como um dos tipsters seguidos (1.176 linhas).
- **Stake em R$** (coluna `Stake`; a coluna `u` é ignorada) — P/L do dashboard em
  reais, como nas contas do Feca / Diogo / Jonathan.
- **Contas verbatim, uma por casa**: as 36 contas da coluna `Conta`
  (`Ikaro`, `Gabi (Suellen)`, `Angela (Tamires)`…) entram como estão, em cada casa
  onde aquela conta apostou → Painel de Contas igual à planilha, custo por linha.
- **Categoria (`Aposta`) = `Múltipla` ou `Outros`.** A planilha NÃO tem coluna de
  mercado e as descrições são curtas e muitas vezes opacas (`Múltipla - Múltipla
  Múltipla` 1.725×, `Mult Betfast` 413×, `Dupla` 330×, `Taborda`, `-` 65×).
  `Múltipla` só onde a descrição PROVA combinação; o resto vai para `Outros` em
  vez de chutar mercado que a planilha não registra.
- **Retorno parcial → odd efetiva (regra de cashout).** Ver abaixo.

── As 67 linhas de retorno parcial ───────────────────────────────────────────

Em 67 linhas o `Green/Red/Half-*` da planilha é o SINAL DO P/L, não o estado das
seleções — 56 delas são bilhete de **sistema** (`X/Y/Z Duplas`, `Tripla`), onde a
odd é a média das linhas e uma dupla certa em três já devolve parte do stake:
`Sydney/Shanxi/Cerro Duplas`, stake 300, marcada `Red` com **−112,50**.

Importar verbatim entregaria `P/L = −300` ali e **+R$ 6.586,54 de lucro fantasma**
no total (derivado 176.406,92 × planilha 169.820,39). Então essas linhas entram
pela regra de cashout do `MASTER_RESULTADO §5.6` (retorno ≠ stake → `W`, com
`Odd = Retorno ÷ Stake`):

    odd = (stake + PL_planilha) / stake        → W   (V quando PL = 0)

O P/L passa a bater centavo a centavo com a planilha. O preço, explícito: nessas
67 linhas a odd exibida é a **efetiva** (0,625 no exemplo — menor que 1, como todo
cashout parcial) e o badge é `W` mesmo onde a planilha dizia `Red`/`Half-*`. As
outras 10.914 linhas fecham `PL = stake × (odd−1) | −stake | −stake/2 | 0` sem
divergência e entram verbatim.

O `_relatorio` refaz essa conferência linha a linha, então erro de normalização
(odd com ponto/vírgula trocados, resultado mal lido) aparece como divergência em
vez de passar calado.

── Esporte: só `Diversos` é traduzido; o resto é a grafia do dono ────────────

A base preserva a grafia herdada de import (é a que o matcher compara) — o mapa
só alinha rótulo com o valor oficial do `MASTER_ESPORTES §7` quando o próprio
sistema já usa outra grafia (`Beisebol` → `Baseball`, `Fut Americano` → `Futebol
Americano`, `Volei` → `Vôlei`, `Tenis` → `Tênis`, `Handball` → `Handebol`,
`Dota` → `E-Sports`), todas MEDIDAS no banco antes de escolher.

`Diversos` (1.776) não é esporte: é a gaveta do dono. Resolvido nesta ordem:

1. descrição prova combinação → **`Múltiplos`** (§2), 1.158 linhas;
2. tipster `eBasket` → **`eBasket`** (o nome do tipster É a modalidade, e as
   descrições são totais de pontos: `TD24 o102.5`), 160 linhas;
3. léxico de **F1** (piloto + termo de sessão: `Piastri t1`, `Hamilton Qualy2`,
   `Norris TOP3`) → **`F1`**, ~174 linhas dos tipsters Deano/Marco;
4. o que sobra → **`Outro`** (§3: sem contexto suficiente). São ~284 linhas
   genuinamente ambíguas (MMA, dardos, e-sports e tênis por set misturados no
   mesmo rótulo) — chutar modalidade ali seria dado errado silencioso.

**O esporte das outras 9.205 linhas fica como o dono escreveu**, inclusive quando
uma múltipla de 3 jogos está marcada `Futebol` em vez de `Múltiplos`: o §2 pediria
a troca, mas eu não sei quantas seleções tem um `Mult Betfast` — reclassificar
milhares de linhas por heurística fraca é pior que preservar o dado dele.

── Casa: grafia canônica, medida no banco ────────────────────────────────────

`casa` é TEXTO em 7 tabelas: cada grafia é uma casa DIFERENTE no sistema. As 37
casas da planilha vêm em CAIXA ALTA (`BET365`) e foram mapeadas para a grafia que
o banco JÁ usa (`select casa, count(*) from bilhetes group by 1`) — nunca
title-casear, que mutila nome e cria conta paralela. `REI DO PITACO` → `Pitaco`
(unificado na s270; a chave velha não casa com nada). 10 casas de baixo volume não
têm correspondência no banco e entram na grafia de marca — o `_relatorio` as lista
uma a uma sob `⚠ SEM CORRESPONDÊNCIA NO BANCO` para o Feca vetar antes do `--go`.

Uso:
    python scripts/import_arrudex_xlsx.py --xlsx "C:\\Users\\Fernando\\Downloads\\2026 - Fernando.xlsx"        # DRY
    python scripts/import_arrudex_xlsx.py --xlsx "C:\\Users\\Fernando\\Downloads\\2026 - Fernando.xlsx" --go   # escreve
"""
import argparse
import asyncio
import datetime as dt
import hashlib
import os
import re
from collections import Counter

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
DONO = 'arrudex'          # username conferido na tabela `usuarios` (não o nome do arquivo)
ABA = 'Apostas'
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}
ODD_OBRIGATORIA = {'W', 'HW'}     # espelha repository._RESULTADOS_ODD_OBRIGATORIA
TOL = 0.05                        # tolerância do casamento de P/L (centavos da planilha)


# ---------- sanitização de texto ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


# ---------- casa ----------
# Grafia canônica MEDIDA no banco antes de escolher (ver docstring). Casa fora do
# mapa entraria verbatim EM CAIXA ALTA — por isso as 37 estão todas aqui.
_CASA_MAP = {
    'BET365': 'Bet365',
    'BETANO': 'Betano',
    'SUPERBET': 'Superbet',
    'PINNACLE': 'Pinnacle',
    'BETFAST': 'Betfast',
    'BETESPORTE': 'BETesporte',
    'BETNACIONAL': 'Betnacional',
    'BETFAIR': 'Betfair',
    'STAKE': 'Stake',
    'FULLTBET': 'Fulltbet',
    'REI DO PITACO': 'Pitaco',          # unificada na s270 — não recriar a grafia velha
    'TIVOBET': 'Tivo',
    'CASA DE APOSTAS': 'Casa de Apostas',
    'BETBRA': 'Betbra',
    'ESPORTIVA': 'Esportiva',
    'KTO': 'KTO',
    'FAZ1BET': 'Faz1bet',               # o banco tem as duas caixas; esta é a majoritária
    'APOSTA1': 'Aposta1',
    'BETBOO': 'Betboo',
    'BETBOOM': 'Betboom',
    'BATEU': 'Bateu',
    'MULTIBET': 'MultiBet',
    'R7': 'R7',
    'ESPORTES DA SORTE': 'Esportes da Sorte',
    'PIXBET': 'PixBet',                 # ⚠ NÃO é Betpix365 — são casas diferentes
    'NOVIBET': 'Novibet',
    'APOSTAGANHA': 'Aposta Ganha',
    'MILHAO': 'Bet do Milhão',          # `milhao.bet.br` — grafia já nos 3 mapas de favicon
    'BINGO': 'BingoBet',                # `bingo.bet.br`; o banco já tem 16 bilhetes nesta grafia
    # Sem correspondência no banco. Os 10 domínios foram confirmados um a um na lista
    # oficial de bets autorizadas da SPA/MF (que registra tudo em CAIXA ALTA, então não
    # serve para a caixa) e a CAIXA saiu do site/canal oficial de cada casa — mesmo
    # método da s272 (`Betvip`, `Suprema Bet`).
    '1XBET': '1xBet',                   # `1xbet.bet.br` — marca global "1xBet"
    'BIG': 'BigBet',                    # `big.bet.br` — o site é "BigBet - A maior do Brasil!"
    'LUVA': 'Luva.Bet',                 # `luva.bet.br` — título do site: "Luva.Bet"
    'BRBET': 'BRBet',                   # `brbet.bet.br` — perfil oficial: "BRBet - A Bet do Brasil"
    # `sporty.bet.br` = SPORTYBET, autorizada e DISTINTA da `Sportingbet`
    # (`sportingbet.bet.br`, 85 bilhetes no banco) — não unificar as duas.
    'SPORTY': 'SportyBet',
    'VIVASORTE': 'Viva Sorte Bet',      # `vivasorte.bet.br` — título do site
    'CBESPORTES': 'CBesportes',         # `cbesportes.bet.br` — grafia da própria casa
    'APOSTOU': 'Apostou',               # `apostou.bet.br` (a institucional é "APOSTOU.COM")
}
# Casas que NÃO tinham nenhum bilhete no banco quando o mapa foi escrito (medido nesta
# sessão) — só para o aviso do relatório; não muda comportamento. `Bet do Milhão` (4
# bilhetes) e `BingoBet` (16) ficam fora porque já existiam nessa grafia exata.
_CASA_NOVA = {'1xBet', 'BigBet', 'Luva.Bet', 'BRBet', 'SportyBet', 'Viva Sorte Bet',
              'CBesportes', 'Apostou'}


def norm_casa(v) -> str:
    bruto = limpa(v)
    return _CASA_MAP.get(bruto.upper(), bruto)


# ---------- combinação (categoria + esporte `Múltiplos`) ----------
# Marca de bilhete combinado. A barra só conta CERCADA DE ESPAÇO: 'Hsieh/Ostapenko'
# é uma dupla de tênis (um par, uma seleção), enquanto 'Union / Plymouth / Limavady'
# são três seleções.
_COMBO = re.compile(
    r'm[uú]ltipl|\bmult\b|\bduplas?\b|\btriplas?\b|\btrixie\b|\bquadrupla\b|\s/\s|\s/|/\s',
    re.I)


def eh_combo(desc: str) -> bool:
    return bool(_COMBO.search(desc))


def categoria(desc: str) -> str:
    """`Múltipla` só quando a descrição prova combinação; o resto é `Outros` —
    a planilha não tem coluna de mercado (decisão do Feca)."""
    return 'Múltipla' if eh_combo(desc) else 'Outros'


def descricao(desc: str) -> str:
    """Verbatim, com as seleções separadas por ' // ' — o ÚNICO separador de
    seleção (regra #19). Só divide na barra cercada de espaço, para não partir
    nome de par ('Hsieh/Ostapenko')."""
    partes = [p.strip() for p in re.split(r'\s+/\s*|\s*/\s+', desc) if p.strip()]
    return ' // '.join(partes) if len(partes) > 1 else desc


# ---------- esporte ----------
# Só alinha rótulo com a grafia que o sistema JÁ usa (medida no banco). O resto
# fica como o dono escreveu.
_ESPORTE_MAP = {
    'Futebol': 'Futebol',
    'Basquete': 'Basquete',
    'Tenis': 'Tênis',
    'Tênis': 'Tênis',
    'Beisebol': 'Baseball',
    'Volei': 'Vôlei',
    'Fut Americano': 'Futebol Americano',
    'Handball': 'Handebol',
    'Hóquei': 'Hóquei',
    'Dota': 'E-Sports',
}

# F1: piloto + termo de sessão. Os dois lados são exigidos para nome de piloto não
# casar com sobrenome de outro esporte.
_F1_PILOTO = re.compile(
    r'\b(piastri|leclerc|hamilton|norris|russel{1,2}|verstappen|gasly|alonso|sainz|'
    r'perez|ocon|hulkenberg|tsunoda|albon|stroll|bottas|zhou|magnussen|ricciardo|'
    r'lawson|bearman|colapinto|antonelli|hadjar|bortoleto|doohan|vespa)\b', re.I)
_F1_TERMO = re.compile(r'\b(qualy\d?|q[1-3]|t[1-3]|top\s?\d+|race|gp|grid|pole|sprint|'
                       r'classifica|corrida)\b', re.I)


def norm_esporte(bruto: str, tipster: str, desc: str) -> str:
    if bruto in _ESPORTE_MAP:
        return _ESPORTE_MAP[bruto]
    if bruto != 'Diversos':
        return bruto                                  # grafia do dono, verbatim
    # `Diversos` é a gaveta do dono — resolvida por ordem de evidência (§2/§3)
    if eh_combo(desc):
        return 'Múltiplos'
    if tipster.lower() == 'ebasket':
        return 'eBasket'
    if _F1_PILOTO.search(desc) and _F1_TERMO.search(desc):
        return 'F1'
    return 'Outro'


# ---------- data / stake / odd ----------
def norm_data(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def _para_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = limpa(v).replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def fmt_stake(v) -> str:
    n = _para_float(v)
    return '' if n is None else f'{n:.2f}'.replace('.', ',')


def _float_str(n: float) -> str:
    """Precisão completa (nunca truncar odd), vírgula decimal."""
    s = repr(round(float(n), 10))
    if s.endswith('.0'):
        s = s[:-2]
    return s.replace(',', '').replace('.', ',')


def norm_odd(v) -> str:
    n = _para_float(v)
    return '' if n is None or n <= 0 else _float_str(n)


# Rótulos da planilha → códigos canônicos do MASTER_RESULTADO. `Aberto` → vazio
# (aposta não liquidada; MASTER_OUTPUT §13.1).
_RESULTADO_MAP = {
    'GREEN': 'W', 'RED': 'L', 'VOID': 'V',
    'HALF-GREEN': 'HW', 'HALF-RED': 'HL', 'ABERTO': '',
}


def norm_resultado(v) -> str:
    r = limpa(v).upper()
    r = _RESULTADO_MAP.get(r, r)
    return r if r in VALID else ''


# ---------- P/L derivado (mesma fórmula do app; nunca uma 2ª) ----------
def pl_derivado(stake: str, odd: str, res: str):
    s = _para_float(stake) or 0.0
    o = _para_float(odd)
    if res == 'L':
        return -s
    if res == 'V':
        return 0.0
    if res == 'HL':
        return -s / 2
    if o is None:
        return None
    if res == 'W':
        return s * (o - 1)
    if res == 'HW':
        return s * (o - 1) / 2
    return None


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        raise SystemExit(f"aba {ABA!r} não encontrada — abas: {wb.sheetnames}")
    ws = wb[ABA]
    # linha de dado = data REAL na col. 0 (descarta cabeçalho, linhas vazias e os
    # painéis de KPI que a planilha mantém da coluna M em diante)
    brutas = [r for r in ws.iter_rows(values_only=True)
              if r and isinstance(r[0], dt.datetime)]
    out = []
    for r in brutas:
        bruto = limpa(r[2])
        tipster = limpa(r[1])
        res = norm_resultado(r[6])
        stake = fmt_stake(r[4])
        odd = norm_odd(r[3])
        row = {
            'data': norm_data(r[0]),
            'esporte': norm_esporte(limpa(r[11]), tipster, bruto),
            'tipster': tipster,
            'casa': norm_casa(r[9]),
            'parceiro': limpa(r[10]),
            'aposta': categoria(bruto),
            'descricao': descricao(bruto),
            'stake': stake,
            'odd': odd,
            'resultado': res,
            '_pl_planilha': _para_float(r[7]),
            '_res_planilha': res,
            '_odd_planilha': odd,
            '_ajustada': False,
            '_bruto': bruto,
        }
        _ajusta_retorno_parcial(row)
        out.append(row)
    return out


def _ajusta_retorno_parcial(row: dict) -> None:
    """Retorno parcial (sistema `X Duplas` e P/L torto) → regra de cashout do
    `MASTER_RESULTADO §5.6`: `W` com `odd = retorno ÷ stake`.

    Sem isso o P/L derivado divergiria da planilha em 67 linhas (+R$ 6.586,54 de
    lucro fantasma). Mexe SÓ onde a conta não fecha — as outras 10.914 linhas
    passam intactas.
    """
    res, pl = row['resultado'], row['_pl_planilha']
    if not res or pl is None:
        return                                   # aberta, ou planilha sem P/L
    stake = _para_float(row['stake']) or 0.0
    if stake <= 0:
        return
    atual = pl_derivado(row['stake'], row['odd'], res)
    if atual is not None and abs(atual - pl) <= TOL:
        return                                   # a conta já fecha
    retorno = stake + pl
    if retorno <= 0:                             # perda total (ou pior) → L
        row['resultado'], row['_ajustada'] = 'L', True
        return
    row['resultado'] = 'W' if abs(pl) > 0.005 else 'V'
    if row['resultado'] == 'W':
        row['odd'] = _float_str(retorno / stake)
    row['_ajustada'] = True


# ---------- assinatura (idêntica a repository._assinatura / _norm_odd) ----------
# ATENÇÃO: `stake` ENTRA no hash. Não copiar `import_lava.py` /
# `import_dashboard_xlsx.py` (versão pré-s133, sem stake). Divergir daqui faz a
# próxima captura gerar assinatura nova e duplicar o histórico inteiro.
def _norm_odd_sig(v: str) -> str:
    try:
        return f"{round(float(v.replace(',', '.')), 2):.2f}"
    except (ValueError, AttributeError):
        return v


def assinaturas(rows: list[dict]) -> list[str]:
    """Assinatura por linha com contador de duplicatas por lote (igual ao app):
    conteúdo idêntico ESCALA o contador, nunca colide."""
    counts: dict[str, int] = {}
    sigs = []
    for r in rows:
        base_raw = "|".join([
            r['casa'], r['parceiro'], r['data'], r['aposta'], r['descricao'],
            r['stake'], _norm_odd_sig(r['odd']),
        ])
        base_sig = hashlib.sha256(base_raw.encode()).hexdigest()[:20]
        cnt = counts.get(base_sig, 0) + 1
        counts[base_sig] = cnt
        raw = base_raw if cnt == 1 else f"{base_raw}|{cnt}"
        sigs.append(hashlib.sha256(raw.encode()).hexdigest()[:20])
    return sigs


def estado_extracao(resultado: str, odd: str) -> str:
    """Espelha repository.estado_extracao: 'resolvida' exige resultado canônico, e
    odd > 0 só onde o P/L depende dela (W/HW)."""
    if resultado not in VALID:
        return 'aberta'
    if resultado in ODD_OBRIGATORIA and (_para_float(odd) or 0) <= 0:
        return 'aberta'
    return 'resolvida'


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict]):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    sigs = assinaturas(rows)
    contas = sorted({(r['casa'], r['parceiro']) for r in rows})

    registros = []
    for r, sig in zip(rows, sigs):
        registros.append((
            DONO, r['casa'], r['parceiro'], sig, None,           # codigo_bilhete
            r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
            r['stake'], r['odd'], r['resultado'] or None,
            estado_extracao(r['resultado'], r['odd']),
            None, None, ORIGEM,                                  # confianca, stake_usd
        ))

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=180)
            try:
                async with conn.transaction():
                    # idempotente: reimportar não acumula (limpa só o que ESTE
                    # import escreveu — captura da extensão tem outra origem)
                    apagadas = await conn.execute(
                        "DELETE FROM bilhetes WHERE dono=$1 AND origem=$2", DONO, ORIGEM)
                    print(f'  [tentativa {tentativa}] limpou import anterior: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    # cadastro das contas: a existência de uma conta não vem do
                    # bilhete (CLAUDE.md) — cada par (casa, conta) da planilha vira
                    # linha no Painel de Contas, pronta para receber custo
                    for casa, nome in contas:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            DONO, casa, nome)
                    # feed ordena por criado_em DESC; num import não existe "envio"
                    # → ancora na data da aposta para sair cronológico
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC,
                                                               id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem=$2
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, DONO, ORIGEM)
                n = await conn.fetchval("SELECT COUNT(*) FROM bilhetes WHERE dono=$1", DONO)
                nc = await conn.fetchval(
                    "SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1", DONO)
                np = await conn.fetchval("SELECT COUNT(*) FROM parceiros WHERE dono=$1", DONO)
                print(f'\nOK — bilhetes dono={DONO}={n} | casas={nc} | contas={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


# ---------- relatório do DRY ----------
def _relatorio(rows: list[dict]):
    print(f'DONO={DONO} | linhas: {len(rows)}')
    datas = sorted(dt.datetime.strptime(r['data'], '%d/%m/%Y') for r in rows)
    print(f'período: {datas[0]:%d/%m/%Y} → {datas[-1]:%d/%m/%Y}')

    print('\n— casa —')
    novas = []
    for k, v in Counter(r['casa'] for r in rows).most_common():
        marca = '  ⚠ SEM CORRESPONDÊNCIA NO BANCO' if k in _CASA_NOVA else ''
        if marca:
            novas.append(k)
        print(f'  {v:6d}  {k}{marca}')
    if novas:
        print(f'  → {len(novas)} casa(s) em grafia de marca (confira antes do --go): '
              f'{", ".join(novas)}')

    print('\n— esporte —')
    for k, v in Counter(r['esporte'] for r in rows).most_common():
        print(f'  {v:6d}  {k}')
    print('\n— categoria —')
    for k, v in Counter(r['aposta'] for r in rows).most_common():
        print(f'  {v:6d}  {k}')
    print('\n— resultado (após o ajuste de retorno parcial) —')
    for k, v in Counter(r['resultado'] or '(aberta)' for r in rows).most_common():
        print(f'  {v:6d}  {k}')
    print('\n— tipster —')
    for k, v in Counter(r['tipster'] for r in rows).most_common():
        print(f'  {v:6d}  {k}')

    contas = sorted({(r['casa'], r['parceiro']) for r in rows})
    print(f'\n— contas: {len(contas)} pares (casa, conta) de '
          f'{len({r["parceiro"] for r in rows})} contas distintas —')
    for k, v in Counter(r['parceiro'] for r in rows).most_common():
        print(f'  {v:6d}  {k}')

    sigs = assinaturas(rows)
    dup = len(sigs) - len(set(sigs))
    print(f'\nassinaturas: {len(set(sigs))} únicas de {len(sigs)} '
          f'({dup} desambiguadas por contador)')
    print('extraction_state:',
          dict(Counter(estado_extracao(r['resultado'], r['odd']) for r in rows)))

    sem_odd = [r for r in rows if not r['odd']]
    print(f'\nsem odd: {len(sem_odd)} linha(s)')
    for r in sem_odd[:20]:
        print(f'    {r["data"]} | {r["resultado"] or "(aberta)":8s} | '
              f'{r["descricao"][:46]}')

    sem_stake = [r for r in rows if (_para_float(r['stake']) or 0) <= 0]
    if sem_stake:
        print(f'\n⚠ {len(sem_stake)} linha(s) com stake 0: gravadas, porém INVISÍVEIS '
              f'no dashboard (dashboard_rows corta stake <= 0).')

    ajust = [r for r in rows if r['_ajustada']]
    print(f'\n=== {len(ajust)} LINHA(S) COM RETORNO PARCIAL → odd efetiva '
          f'(MASTER_RESULTADO §5.6) ===')
    for r in ajust:
        print(f'  {r["data"]} {r["_res_planilha"]:2s}→{r["resultado"]:1s} '
              f'stake={r["stake"]:>9s} odd {r["_odd_planilha"] or "—"}→{r["odd"] or "—"} '
              f'PL_plan={r["_pl_planilha"]:>9.2f}  {r["_bruto"][:38]}')

    # conferência final: o P/L derivado tem de bater com a coluna PL R$ da planilha
    liquidadas = [r for r in rows if r['resultado']]
    div = [r for r in liquidadas
           if (v := pl_derivado(r['stake'], r['odd'], r['resultado'])) is None
           or abs(v - (r['_pl_planilha'] or 0)) > TOL]
    pl_der = sum(v for r in liquidadas
                 if (v := pl_derivado(r['stake'], r['odd'], r['resultado'])) is not None)
    pl_plan = sum(r['_pl_planilha'] or 0 for r in liquidadas)
    turnover = sum(_para_float(r['stake']) or 0 for r in rows)
    print(f'\nP/L derivado × coluna PL R$ da planilha: {len(div)} divergência(s)')
    for r in div[:15]:
        print(f'    {r["data"]} | {r["resultado"]} | {r["descricao"][:40]} | '
              f'planilha={r["_pl_planilha"]}')
    print(f'\nturnover:      R$ {turnover:>14,.2f}')
    print(f'P/L planilha:  R$ {pl_plan:>14,.2f}')
    print(f'P/L derivado:  R$ {pl_der:>14,.2f}   (diferença: {pl_der - pl_plan:+,.2f})')

    print('\n=== 20 AMOSTRAS (Data|Esporte|Tipster|Casa|Conta|Aposta|Descrição|Stake|Odd|Res) ===')
    step = max(1, len(rows) // 20)
    for r in rows[::step][:20]:
        print(' | '.join([r['data'], r['esporte'], r['tipster'], r['casa'], r['parceiro'],
                          r['aposta'], r['descricao'][:40], r['stake'], r['odd'],
                          r['resultado'] or '—']))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows))


if __name__ == '__main__':
    main()
