# -*- coding: utf-8 -*-
"""Importa base já no LAYOUT DASHBOARD (.xlsx, aba 'Dados').

Diferente de import_apostas_xlsx.py (export cru de 15 col do Diogo/Primo): aqui a
planilha já sai no formato de saída do projeto — 10 colunas do MASTER_OUTPUT +
Lucro/Usuário/Duplicata. Ordem da aba 'Dados':

    0 Data | 1 Esporte | 2 Tipster | 3 Casa | 4 Parceiro | 5 Aposta |
    6 Descrição | 7 Stake | 8 Odd | 9 Resultado | 10 Lucro | 11 Usuário | 12 Duplicata

Diferenças que exigem de-para mesmo assim:
  • Resultado já vem W/L/V/HW/HL — só normaliza caixa (l→L) e valida.
  • Coluna 'Aposta' é uma MISTURA (tipo de bilhete + mercado cru): 'Simples',
    'Aumentada', 'Criar Aposta', 'Bet Builder', 'Total', 'Pontos'… A categoria
    canônica (27) sai do motor de keywords aplicado a Aposta + Descrição.
  • Bilhetes combinados (Aumentada / Criar Aposta / Bet Builder / Duplas / Treble)
    ou descrição multi-perna (' // ') → 'Múltipla' (decisão do Feca).
  • Esporte/Casa ao canon; Parceiro REAL é mantido (só funde duplicata de caixa).
  • Lucro (derivado) e Usuário (email) são descartados. P/L é derivado no app.

Modelo de conta: Jonathan é DONO sem operadores. base sob dono='Jonathan'.

Uso:
    python scripts/import_dashboard_xlsx.py --dono Jonathan --xlsx "C:\\...\\base.xlsx"        # DRY
    python scripts/import_dashboard_xlsx.py --dono Jonathan --xlsx "C:\\...\\base.xlsx" --go     # escreve
"""
import argparse
import asyncio
import hashlib
import os
import re
from collections import Counter

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}


# ---------- sanitização de texto ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')       # C0 + DEL (inclui \n \r \t \x0c)
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')      # escape literal do openpyxl (_x000C_)


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


# ---------- casa → canon (espelha _casa_display + data.js, com folds desta base) ----------
_CASA_DISPLAY = {
    "BET365": "Bet365", "BETANO": "Betano", "BETFAIR": "Betfair",
    "BETNACIONAL": "Betnacional", "BOLSADEAPOSTA": "Bolsa de Aposta",
    "BOLSADEAPOSTAS": "Bolsa de Aposta",          # fold plural → canon
    "JOGODEOURO": "Jogo de Ouro", "KINGPANDA": "KingPanda", "KTO": "KTO",
    "LOTTU": "Lottu", "PINNACLE": "Pinnacle", "POLYMARKET": "Polymarket",
    "SUPERBET": "Superbet", "VITORIABET": "Vitória Bet",
    "NOVIBET": "Novibet", "BINGOPLUS": "BingoPlus", "VAIDEBET": "VaiDeBet",
    "BETMGM": "BetMGM", "ESPORTESDASORTE": "Esportes da Sorte",
    "SPORTINGBET": "SportingBet", "BETESPORTE": "BETesporte", "BETFAST": "Betfast",
    # casas próprias desta base
    "ESPORTIVABET": "Esportiva Bet", "BETBRA": "BetBra", "PAGOL": "Pagol",
    "FULLBET": "Fullbet", "FULLTBET": "Fullbet", "MULTIBET": "Multibet",
    "APOSTAGANHA": "Aposta Ganha", "PIXBET": "Pixbet", "MATCHBOOK": "Matchbook",
    "BETBOOM": "Betboom", "FAZ1BE": "Faz1be", "CASADEAPOSTAS": "Casa de Apostas",
}


def norm_casa(c: str) -> str:
    key = limpa(c).upper().replace(' ', '')
    return _CASA_DISPLAY.get(key, limpa(c).title())


# ---------- esporte → canon ----------
_ESPORTE_MAP = {
    'futebol': 'Futebol', 'futebol feminino': 'Futebol',
    'mls': 'Futebol', 'la liga': 'Futebol', 'championship': 'Futebol',
    'veikkausliiga': 'Futebol',
    'tênis': 'Tênis', 'tenis': 'Tênis',
    'tênis de mesa': 'Tênis de Mesa', 'tenis de mesa': 'Tênis de Mesa',
    'basquete': 'Basquete', 'basquete 3x3': 'Basquete',
    'nba': 'Basquete', 'wnba': 'Basquete',
    'e-sports': 'E-Sports', 'esports': 'E-Sports', 'e-sport': 'E-Sports',
    'e-football': 'E-Sports', 'e-football ': 'E-Sports',
    'dardos': 'Dardos', 'vôlei': 'Vôlei', 'volei': 'Vôlei', 'voleibol': 'Vôlei',
    'beisebol': 'Baseball', 'baseball': 'Baseball',
    'handebol': 'Handebol', 'badminton': 'Badminton',
    'mma': 'MMA', 'luta': 'MMA',
    'críquete': 'Críquete', 'criquete': 'Críquete', 'cricket': 'Críquete',
    'hóquei no gelo': 'Hóquei', 'hoquei no gelo': 'Hóquei', 'nhl': 'Hóquei',
    'fórmula 1': 'Fórmula 1', 'formula 1': 'Fórmula 1', 'f1': 'Fórmula 1',
    'automobilismo': 'Fórmula 1',
    'rugby': 'Rugby',
    'multiplos': 'Múltiplos', 'múltiplos': 'Múltiplos', 'vários': 'Múltiplos',
    'varios': 'Múltiplos',
    'outro': 'Outros', 'outros': 'Outros', 'outros esportes': 'Outros',
}


def norm_esporte(esp: str) -> str:
    e = limpa(esp)
    return _ESPORTE_MAP.get(e.lower(), e)


# ---------- fold de caixa (tipster e parceiro): "se tá assim, tá assim" ----------
def construir_fold(brutos: list[str]) -> dict[str, str]:
    freq = Counter(t for t in brutos if t)
    canon: dict[str, str] = {}
    for lk in {t.lower() for t in freq}:
        variantes = [(freq[t], t) for t in freq if t.lower() == lk]
        canon[lk] = max(variantes)[1]          # mais frequente vence (empate → maior string)
    return canon


# ---------- resultado: já vem canônico, só normaliza caixa e valida ----------
def norm_resultado(v: str) -> str:
    r = limpa(v).upper()
    return r if r in VALID else ''


# ---------- categoria: motor de keywords sobre Aposta + Descrição ----------
# bilhete combinado → Múltipla (label OU multi-perna na descrição)
_MULTI_LABELS = {
    'múltipla', 'multipla', 'multiplas', 'múltiplas', 'dupla', 'duplas', 'duplass',
    'tripla', 'triplas', 'treble', 'criar aposta', 'bet builder', 'aumentada',
}


def eh_multipla(aposta_label: str, esporte: str, descricao: str) -> bool:
    if limpa(aposta_label).lower() in _MULTI_LABELS:
        return True
    if ' // ' in f' {limpa(descricao)} ':          # pernas separadas por //
        return True
    return norm_esporte(esporte) == 'Múltiplos'


# rótulos que já SÃO categoria canônica (27) → respeitar a classificação do tipster
# em vez de re-inferir da descrição. Sem isso, ML puro ("Fulano [Fulano v Beltrano]")
# cai em Outros porque a descrição é só um nome, sem palavra-chave.
_CANON_KEEP = {
    'ML', 'Handicap', 'Player Props', 'Cartões', 'Games', 'Sets', 'Team Props',
    'Desarmes', 'Anytime', 'DNB', 'Chutes', 'Chutes no Gol', 'Gols', 'Escanteios',
    'Dupla Chance', 'Ambas Marcam', 'Corridas', 'Assistência', 'H2H',
    'Double-Double', 'Triplo-Duplo', 'Impedimentos', 'Jardas', 'Legs',
}


def categoria(aposta_label: str, esporte: str, descricao: str) -> str:
    if eh_multipla(aposta_label, esporte, descricao):
        return 'Múltipla'
    if limpa(aposta_label) in _CANON_KEEP:         # rótulo já canônico → respeita
        return limpa(aposta_label)
    esp = norm_esporte(esporte)
    # rótulo (mercado) + descrição alimentam o matcher — cobre 'Simples' (categoria
    # sai da descrição) e mercados crus ('Money Line', 'Empate Anula', 'Total'…)
    txt = limpa(aposta_label).lower() + ' || ' + limpa(descricao).lower()

    def has(*ks):
        return any(k in txt for k in ks)

    # Fórmula 1: comparativo/vencedor/prop de piloto
    if esp == 'Fórmula 1':
        if has('comparativ', 'head to head', 'h2h'):   return 'H2H'
        if has('vencedor', 'vencer'):                   return 'ML'
        return 'Player Props'

    # específicos primeiro (corner-handicap cai em Escanteios, não Handicap)
    if has('escanteio', 'corner', 'cantos'):                    return 'Escanteios'
    if has('cartõe', 'cartoe', 'cartão', 'cartao', 'card'):     return 'Cartões'
    if has('impedimento', 'offside'):                           return 'Impedimentos'
    if has('desarme', 'tackle'):                                return 'Desarmes'
    if has('no alvo', 'shots on target', 'chutes no gol', 'chutes ao gol',
           'finalizações no gol'):                              return 'Chutes no Gol'
    if has('chute', 'finalizaç', 'finalizacoes', 'shots'):      return 'Chutes'
    if has('jarda', 'yard'):                                    return 'Jardas'
    if has('ambas', 'ambos marcam', 'ambos os times marcam', 'btts',
           'both teams'):                                       return 'Ambas Marcam'
    if has('dupla chance', 'chance dupla', 'double chance'):    return 'Dupla Chance'
    if has('empate anula', 'draw no bet', 'dnb'):               return 'DNB'
    if esp == 'Dardos' and has('180'):                          return 'Legs'
    if has('best of', 'first to', 'legs', 'leg '):              return 'Legs'
    if has('hat-trick', 'hat trick', 'marcador', 'para marcar', 'de cabeça',
           'a qualquer momento', 'anytime'):                    return 'Anytime'
    # estatísticas de Baseball
    if has('total de bases', 'strikeout', 'lançador', 'lancador', 'home run',
           'corrida', 'runs', 'rbi', 'baseball'):               return 'Corridas'
    if has('marca nos dois tempos', 'equipe marca', 'nos dois tempos', 'team total',
           'equipe -', 'equipe total'):                         return 'Team Props'
    # basquete: totais/pontos/rebotes = estatística individual (Player Props)
    if esp == 'Basquete' and has('pontos', 'rebote', 'toco', 'roubo', 'arremesso',
                                 'cestas', 'três pontos', 'threes', 'total'):
        return 'Player Props'
    if has('assistência', 'assistencia', 'assist'):             return 'Assistência'
    if has('duplo-duplo', 'double-double', 'double double'):    return 'Double-Double'
    if has('triplo-duplo', 'triple-double'):                    return 'Triplo-Duplo'
    if has('total de jogos', 'games', 'game ', 'tie break'):    return 'Games'
    if has('sets', 'set ', 'set betting'):                      return 'Sets'
    if has('total de gol', 'gols', 'gol ', 'over/under', 'totais do jogo',
           'total do jogo'):                                    return 'Gols'
    if has('handicap', 'spread', 'linha'):                      return 'Handicap'
    if has('cestas', 'arremesso', '3 convertidas', 'três pontos', 'defesas',
           'goleiro', 'falta', 'pontos', 'rebote', 'roubos', 'jogador',
           'player'):                                           return 'Player Props'
    if has('moneyline', 'money line', 'vencedor', 'resultado', '1x2', 'para vencer',
           'match winner', 'partida', 'vencer', 'para ganhar', 'campeão'):
        return 'ML'
    if has('total', 'over', 'under', 'mais de', 'menos de'):    return 'Outros'
    return 'Outros'


# ---------- data / stake / odd ----------
def norm_data(v) -> str:
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def norm_num(v) -> str:
    if v is None:
        return ''
    return limpa(v).replace('.', ',')


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if 'Dados' not in wb.sheetnames:
        raise SystemExit(f"aba 'Dados' não encontrada — abas: {wb.sheetnames}")
    ws = wb['Dados']
    brutas = [r for r in ws.iter_rows(min_row=2, values_only=True)
              if r and r[0] is not None]          # descarta linha sem data
    fold_tip = construir_fold([limpa(r[2]) for r in brutas])
    fold_par = construir_fold([limpa(r[4]) for r in brutas])
    out = []
    for r in brutas:
        tip, par = limpa(r[2]), limpa(r[4])
        # parceiro que na verdade é nome de CASA (erro de digitação da origem) → 'revisar'
        if par and par.upper().replace(' ', '') in _CASA_DISPLAY:
            parceiro = 'revisar'
        else:
            parceiro = fold_par.get(par.lower(), par) or 'Padrão'
        out.append({
            'data': norm_data(r[0]),
            'esporte': norm_esporte(r[1]),
            'tipster': fold_tip.get(tip.lower(), tip),
            'casa': norm_casa(r[3]),
            'parceiro': parceiro,
            'aposta': categoria(r[5], r[1], r[6]),
            'descricao': limpa(r[6]),
            'stake': norm_num(r[7]),
            'odd': norm_num(r[8]),
            'resultado': norm_resultado(r[9]),
        })
    return out


# ---------- assinatura (idêntica a repository._assinatura / _norm_odd) ----------
def _norm_odd(v: str) -> str:
    try:
        return f"{round(float(v.replace(',', '.')), 2):.2f}"
    except (ValueError, AttributeError):
        return v


def assinaturas(rows: list[dict]) -> list[str]:
    counts: dict[str, int] = {}
    sigs = []
    for r in rows:
        base_raw = "|".join([
            r['casa'], r['parceiro'], r['data'], r['aposta'], r['descricao'],
            _norm_odd(r['odd']),
        ])
        base_sig = hashlib.sha256(base_raw.encode()).hexdigest()[:20]
        cnt = counts.get(base_sig, 0) + 1
        counts[base_sig] = cnt
        raw = base_raw if cnt == 1 else f"{base_raw}|{cnt}"
        sigs.append(hashlib.sha256(raw.encode()).hexdigest()[:20])
    return sigs


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict], dono: str):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    sigs = assinaturas(rows)
    pares = sorted(set((r['casa'], r['parceiro']) for r in rows))

    registros = []
    for r, sig in zip(rows, sigs):
        estado = 'resolvida' if r['resultado'] in VALID else 'aberta'
        registros.append((
            dono, r['casa'], r['parceiro'], sig, None,           # codigo_bilhete
            r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
            r['stake'], r['odd'], r['resultado'] or None, estado,
            None, None, ORIGEM,                                  # confianca, stake_usd, origem
        ))

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=120)
            try:
                async with conn.transaction():
                    apagadas = await conn.execute(
                        "DELETE FROM bilhetes WHERE dono=$1 AND origem='import'", dono)
                    print(f'  [tentativa {tentativa}] limpou parciais: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    for casa, parceiro in pares:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            dono, casa, parceiro)
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC, id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem='import'
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, dono)
                n = await conn.fetchval("SELECT COUNT(*) FROM bilhetes WHERE dono=$1", dono)
                nc = await conn.fetchval(
                    "SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1", dono)
                np = await conn.fetchval("SELECT COUNT(*) FROM parceiros WHERE dono=$1", dono)
                print(f'\nOK — bilhetes dono={dono}={n} | casas sidebar={nc} | parceiros={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


def _relatorio(rows: list[dict], dono: str):
    print(f'DONO={dono} — linhas a importar: {len(rows)}')
    print('\ncasas:', dict(Counter(r['casa'] for r in rows).most_common()))
    print('\nesporte:', dict(Counter(r['esporte'] for r in rows).most_common()))
    print('\ncategoria (aposta):', dict(Counter(r['aposta'] for r in rows).most_common()))
    print('\nresultado:', dict(Counter(r['resultado'] or '(aberta)' for r in rows).most_common()))
    print('\nparceiro:', dict(Counter(r['parceiro'] for r in rows).most_common()))
    print('\ntipster (top20):', dict(Counter(r['tipster'] for r in rows).most_common(20)))
    sigs = assinaturas(rows)
    dup = len(sigs) - len(set(sigs))
    print(f'\nassinaturas: {len(set(sigs))} únicas de {len(sigs)} ({dup} desambiguadas por contador)')
    print('\n=== 14 AMOSTRAS (Data | Esporte | Tipster | Casa | Parceiro | Aposta | Descrição | Stake | Odd | Resultado) ===')
    step = max(1, len(rows) // 14)
    for r in rows[::step][:14]:
        print(' | '.join([
            r['data'], r['esporte'], r['tipster'], r['casa'], r['parceiro'],
            r['aposta'], r['descricao'][:46], r['stake'], r['odd'], r['resultado'] or '—',
        ]))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dono', required=True)
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows, args.dono)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows, args.dono))


if __name__ == '__main__':
    main()
