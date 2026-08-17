# -*- coding: utf-8 -*-
"""Importa a base do tipster PassaTips VIP (`PassaTips VIP.xlsx`) para
dono='passapano'.

⚠️ MARCA ≠ USERNAME. A marca é `PassaTips VIP`; o username com que ele se
cadastrou no site é `passapano`. O `dono` é SEMPRE o username — conferido na
tabela `usuarios` antes do import (`status='ativo'`, hash de 60 caracteres,
e-mail `joaopmoss7@gmail.com`, criado em 17/08/2026), não deduzido do nome do
arquivo. É a regra da s260: com o `dono` errado o isolamento falha em SILÊNCIO —
nenhum erro, só tela vazia para o usuário certo. A ponte entre os dois nomes é o
registro `TIPSTERS_PUBLICOS` (`app/main.py`), onde o slug é a marca.

A conta foi criada pelo PRÓPRIO usuário no site (Fase 2) e aprovada pelo Feca.
**Não existe env var nem linha em `app/auth.py` para ela.** A base estava vazia
(0 bilhetes) antes deste import.

── Layout da planilha ────────────────────────────────────────────────────────

Três abas (`JUNHO`, `JULHO`, `AGOSTO`), 02/06 → 17/08/2026. Em todas, o
cabeçalho está na linha 15 e os dados começam na 16. **Só as colunas B..I são
apostas**; da M em diante há um painel de "LUCRO DIÁRIO / RÓI DIÁRIO" por data,
que é tabela auxiliar da própria planilha e é IGNORADO (lê-lo dobraria a coluna
Data e criaria linhas fantasma):

    B Data | C Casa de Aposta | D Esporte | E Descrição | F Odd |
    G Resultado | H Aposta (stake em unidades) | I Lucro

A linha de dado é reconhecida por `Data` REAL na coluna B — mesmo critério do
`import_reidocriquete_xlsx.py` e do `import_fleury_xlsx.py`.

**A planilha é consistente e isso foi MEDIDO, não suposto:** o `Lucro` bate com
`stake × (odd − 1)` nas 382 greens e com `−stake` nas 498 reds — **911 de 911,
zero divergências**. O `_relatorio` refaz essa conta linha a linha, então
qualquer erro de normalização (odd com ponto/vírgula trocados, resultado mal
lido) aparece como divergência em vez de passar calado.

── Esporte: 3 rótulos da coluna NÃO são esporte ──────────────────────────────

`Múltiplas` (79), `AllSports` (34) e `Aumentada` (41) são tipo de bilhete, não
modalidade. Decisão do Feca:

- `Múltiplas` / `AllSports` → esporte **`Múltiplos`**, o valor oficial do
  `MASTER_ESPORTES §7` para bilhete que mistura esportes ou acumula seleções de
  jogos diferentes. A coluna `Aposta` segue `Múltipla` (o §7 é explícito: o
  esporte especial não substitui a categoria).
- `Aumentada` → esporte **`Futebol`**. É o "aumento de odd" da Betano (bet
  builder do mesmo jogo), não uma modalidade — e as **41 de 41** foram medidas
  como futebol (cantos, cartões, gols, ambas marcam, impedimentos; zero
  ocorrências de games/aces/hits/pontos/kills). Categoria `Múltipla`, porque bet
  builder é Múltipla mesmo com tudo do mesmo jogo (`MASTER_APOSTAS`, s269).

⚠️ SETE esportes NÃO são canônicos no `MASTER_ESPORTES §7`: `Polo Aquático` (6),
`Críquete` (2), `Vôlei de Praia` (3), `Futebol de Praia` (2), `Boxe` (1),
`Tênis de Mesa` (1), `Futsal` (1). Eles entram na grafia que o sistema JÁ
conhece (o precedente do `CLAUDE.md`: a base preserva a grafia herdada de import
e a rota `/taxonomia` serve a UNIÃO do canônico com a base do dono, então os
menus dele mostram os sete). **Não** mexi no MASTER — criar esporte é a regra de
propagação, que é mudança própria com aprovação.

Duas normalizações de grafia dentro da própria planilha, para não nascer esporte
gêmeo na base do dono: `Vôlei de praia`/`Vôlei de Praia` → uma só, e
`Tenis de Mesa` → `Tênis de Mesa` (a grafia citada no `CLAUDE.md`). `Beisebol` →
`Baseball` e `LOL`/`CS`/`Dota` → `E-Sports`, que são os valores oficiais do §7.

── Categoria (coluna `Aposta`) ───────────────────────────────────────────────

Derivada da DESCRIÇÃO, seguindo o princípio do `MASTER_APOSTAS §1`: **a
categoria registra o OBJETO da aposta, não o tipo de mercado.** Por isso
`HA cantos Vardar +2.5` é `Escanteios` (não `Handicap`) e `HA games Dylan +2.5`
é `Games` — o handicap é a forma, o objeto é o que conta. `Handicap` e `ML` só
entram quando o objeto é o RESULTADO e nenhum outro objeto foi nomeado.

A ordem das regras é a do `MASTER_APOSTAS §2` (categoria específica → regra por
esporte → Player Props → Outros) e duas precedências são load-bearing:

- **`Pontos` antes de `Sets`** — no Vôlei a UNIDADE CONTADA decide (`§7`):
  `1set under 45.5 pontos` conta pontos, não sets. Inverter a ordem manda 7
  linhas de vôlei para a gaveta errada, e o sintoma é mudo.
- **`Games` antes de `Sets`** — `1set under 8.5 games` conta games.

Resultado medido nas 911 linhas: **1 única linha em `Outros`** (`Over 10.5
ensaios`, rugby — o MASTER não tem categoria para tries).

As 9 linhas de futebol com `faltas` entram em **`Faltas`**, categoria criada no
`MASTER_APOSTAS §3` na mesma sessão (s272) com a regra de propagação inteira
(§3, §4, §5, §6, §7, §9) e o repontamento do `§9` de três casas que mandavam
faltas para a gaveta errada: `CASA_BETANO` e `CASA_BETFAST` (`Outros ⚠️`) e
`CASA_KTO` (`Player Props`).

Estas 9 linhas **não** foram a evidência que sustentou a categoria — a régua do
`§8` (`Outros` é último recurso) e a do `§2` (nunca `Player Props` quando há
categoria específica) já vinham sendo forçadas em três casas, e o Feedback da
`CASA_BETFAST` registrava faltas em **11 pernas numa conta só, mais que `Gols`**.

── Descrição ─────────────────────────────────────────────────────────────────

VERBATIM da planilha, só com limpeza de espaço e caractere de controle. O
`MASTER_DESCRICAO §2` pede `Entidade - Mercado [Confronto]`, mas esta fonte não
registra o confronto em coluna nenhuma — e o `§1` é explícito: **nunca inventar
informação**. O confronto existe nos prints do canal do Telegram; recuperá-lo
exigiria visão sobre 859 imagens e é frente própria, não deste import.

── Casas ─────────────────────────────────────────────────────────────────────

`casa` é TEXTO em 7 tabelas: cada grafia é uma casa DIFERENTE, com conta, KPI,
filtro e favicon próprios. As grafias abaixo foram MEDIDAS contra as 59 do banco
antes de escolher — nenhuma foi inventada nem title-caseada:

- `Bet365`, `Betano`, `Betnacional`, `Novibet` → já canônicas, batem exato.
- `Estrela` → **`Estrela Bet`**, que é a grafia do banco (15 bilhetes) e a dos
  três mapas de favicon (`index.html`, `data.js`, `inicio.html`).
- `Betanacional` (4 linhas) → **`Betnacional`**. É erro de digitação da mesma
  casa, não uma casa nova; deixá-lo passar criaria uma 60ª grafia gêmea.
- `Betvip` (27) e `Supremabet` (4) → **casas NOVAS** (não existem em nenhuma das
  59 grafias). A caixa foi conferida no site de cada uma antes de gravar:
  `betvip.bet.br` se apresenta como **`Betvip`** e `suprema.bet.br` como
  **`Suprema Bet`** ("operado pela SUPREMA BET LTDA"). Decisão do Feca: nome de
  marca do site.

── Numeração ─────────────────────────────────────────────────────────────────

Código `PT<aaaamm>-<n>`, na família do `SC` (Só Chutes), `ZE` (Zora) e `RC` (Rei
do Criquete). Numeração MENSAL: reinicia em 1 a cada mês, na ordem das linhas da
planilha. As premissas foram medidas: **1 aba = 1 mês** (JUNHO 06/2026, JULHO
07/2026, AGOSTO 08/2026) e **zero linhas fora de ordem cronológica** dentro de
cada aba — o script ABORTA se a primeira cair.

Isso não é cosmético: `repository._assinatura` com código é
`ID|casa|parceiro|codigo` — o CONTEÚDO não entra no hash. Numerar é o que faz o
bot, ao planilhar a próxima aposta do canal, casar com a linha certa em vez de
duplicar o histórico.

── Decisões ──────────────────────────────────────────────────────────────────

- **Dono solo** — sem OPERADORES, sem dedup cruzada.
- **Tipster** = `PassaTips VIP` (nome de marca) em todas as linhas.
- **Stake em UNIDADES** (0,25–4,00; a planilha declara "Aposta média 1,28"), 1u =
  1 — o P/L do dashboard é o P/L em unidades. Mesma decisão do `SoChutes`, do
  `Fleury` e do `Rei do Criquete`.
- **Uma conta `Padrão` por casa** — 7 linhas no Painel de Contas, cada uma com
  custo próprio.
- Resultado: `green` → W · `red` → L · `void` → V · vazio → **aberta** (3 linhas:
  duas de 31/07 na Betvip e uma de 17/08 na Betnacional).

Uso:
    python scripts/import_passatips_xlsx.py --xlsx "C:\\...\\PassaTips VIP.xlsx"        # DRY
    python scripts/import_passatips_xlsx.py --xlsx "C:\\...\\PassaTips VIP.xlsx" --go   # escreve
"""
import argparse
import asyncio
import datetime as dt
import hashlib
import os
import re
import unicodedata
from collections import Counter, defaultdict

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
DONO = 'passapano'            # username conferido em `usuarios` (marca: PassaTips VIP)
PARCEIRO = 'Padrão'
TIPSTER = 'PassaTips VIP'
PREFIXO = 'PT'                # código PT<aaaamm>-<n>, família do SC/ZE/RC
ABAS = ('JUNHO', 'JULHO', 'AGOSTO')   # ordem cronológica (o arquivo traz ao contrário)
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}
LINHA_DADOS = 16              # cabeçalho na 15; da 16 em diante são apostas
COL_INI, COL_FIM = 2, 9       # B..I — da M em diante é painel diário, ignorado


# ---------- sanitização de texto ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


def _chave(s) -> str:
    """minúscula sem acento — para casar rótulo digitado à mão (`Tenis`/`Tênis`,
    `Vôlei de praia`/`Vôlei de Praia`)."""
    s = unicodedata.normalize('NFKD', limpa(s).lower())
    return ''.join(c for c in s if not unicodedata.combining(c))


def _k(s) -> str:
    """_chave sem espaço — para casar nome de casa/esporte com espaçamento livre."""
    return _chave(s).replace(' ', '')


# ---------- casa ----------
# Grafias MEDIDAS contra as 59 do banco. Casa fora deste mapa entra VERBATIM
# (nunca title-casear: mutila nome e cria conta paralela) — e o relatório do DRY
# avisa, para ninguém descobrir uma 60ª grafia depois de gravar.
_CASA_MAP = {
    'bet365': 'Bet365',
    'betano': 'Betano',
    'betnacional': 'Betnacional',
    'betanacional': 'Betnacional',   # erro de digitação da MESMA casa (4 linhas)
    'novibet': 'Novibet',
    'estrela': 'Estrela Bet',        # grafia do banco e dos 3 mapas de favicon
    'estrelabet': 'Estrela Bet',
    'betvip': 'Betvip',              # casa NOVA — caixa conferida em betvip.bet.br
    'supremabet': 'Suprema Bet',     # casa NOVA — "operado pela SUPREMA BET LTDA"
    'suprema': 'Suprema Bet',
}
CASAS_NOVAS = {'Betvip', 'Suprema Bet'}   # não existiam nas 59 grafias do banco


def norm_casa(v) -> str:
    bruto = limpa(v)
    return _CASA_MAP.get(_k(bruto), bruto)


# ---------- esporte ----------
# `Múltiplas`/`AllSports` → `Múltiplos` (valor especial do §7). `Aumentada` →
# `Futebol` (41/41 medidas como futebol). Ver docstring.
_ESPORTE_MAP = {
    'futebol': 'Futebol',
    'multiplas': 'Múltiplos',
    'allsports': 'Múltiplos',
    'aumentada': 'Futebol',
    'tenis': 'Tênis',
    'basquete': 'Basquete',
    'beisebol': 'Baseball',
    'volei': 'Vôlei',
    'futebolamericano': 'Futebol Americano',
    'handebol': 'Handebol',
    'hoquei': 'Hóquei',
    'dardos': 'Dardos',
    'rugby': 'Rugby',
    'badminton': 'Badminton',
    # E-Sports: o §7 tem um valor só; o jogo específico vive na descrição.
    'lol': 'E-Sports',
    'cs': 'E-Sports',
    'dota': 'E-Sports',
    # NÃO-CANÔNICOS (não existem no MASTER_ESPORTES §7) — grafia que o sistema
    # já conhece, para não nascer esporte gêmeo na base do dono.
    'poloaquatico': 'Polo Aquático',
    'cricket': 'Críquete',
    'voleidepraia': 'Vôlei de Praia',
    'futeboldepraia': 'Futebol de Praia',
    'boxe': 'Boxe',
    'tenisdemesa': 'Tênis de Mesa',
    'futsal': 'Futsal',
}
NAO_CANONICOS = {'Polo Aquático', 'Críquete', 'Vôlei de Praia', 'Futebol de Praia',
                 'Boxe', 'Tênis de Mesa', 'Futsal'}
# Rótulos da coluna Esporte que são TIPO DE BILHETE, não modalidade.
PSEUDO_ESPORTES = {'multiplas', 'allsports', 'aumentada'}


def norm_esporte(v) -> str:
    return _ESPORTE_MAP.get(_k(v), limpa(v) or 'Outro')


# ---------- categoria ----------
# MASTER_APOSTAS §1: a categoria registra o OBJETO da aposta. O tipo de mercado
# (handicap / total / comparativo) NÃO muda a categoria. Ver docstring.
_RE_COMBO = re.compile(r'^\s*(dupla|tripla|quadrupla|quintupla|multipla)s?\b')
_RE_DC = re.compile(r'\bdc\b|dupla chance')
_RE_FALTAS_TENIS = re.compile(r'duplas? faltas?')
# "tem número/linha" = é mercado de total; sem isso, nome de time solto é resultado.
_RE_TOTAL = re.compile(r'\b(over|under|mais de|menos de)\b|[+-]?\d+[.,]\d')

# Objeto padrão do total quando a descrição NÃO nomeia o objeto
# ("Under 87.5 HT" no basquete são pontos; "Redblacks HT" é resultado).
_TOTAL_POR_ESPORTE = {
    'Basquete': 'Pontos', 'Futebol Americano': 'Pontos', 'Vôlei': 'Pontos',
    'Vôlei de Praia': 'Pontos', 'Badminton': 'Pontos', 'Tênis de Mesa': 'Pontos',
    'Futebol': 'Gols', 'Futsal': 'Gols', 'Handebol': 'Gols', 'Hóquei': 'Gols',
    'Polo Aquático': 'Gols', 'Futebol de Praia': 'Gols',
    'Tênis': 'Games', 'Baseball': 'Corridas', 'E-Sports': 'E-Sports Props',
    'Boxe': 'Rounds',
}


def norm_categoria(esporte_bruto, esporte, descricao) -> str:
    d = _chave(descricao)

    # 1. Múltipla — pelo pseudo-esporte da planilha ou pelo prefixo do cupom.
    #    `DC …` e `duplas faltas` NÃO são combo, apesar do "dupla".
    if _k(esporte_bruto) in PSEUDO_ESPORTES:
        return 'Múltipla'
    if _RE_COMBO.match(d) and not _RE_DC.search(d) and not _RE_FALTAS_TENIS.search(d):
        return 'Múltipla'

    # 2. OBJETO da aposta — vence sempre o tipo de mercado (§1).
    if re.search(r'escanteio|\bcantos?\b', d):
        return 'Escanteios'
    if re.search(r'\bcart(ao|oes)\b|expuls|vermelho', d):
        return 'Cartões'
    if re.search(r'impedimento|offside', d):
        return 'Impedimentos'
    if re.search(r'\bsots?\b|chutes? no gol|no alvo', d):
        return 'Chutes no Gol'
    if re.search(r'\bchutes?\b|finaliza', d):
        return 'Chutes'
    if re.search(r'\bdesarmes?\b|\btackles?\b', d):
        return 'Desarmes'
    if re.search(r'\bgols?\b', d):
        return 'Gols'
    if re.search(r'\bgames?\b', d):          # ANTES de Sets: `1set under 8.5 games`
        return 'Games'
    if re.search(r'\bkills?\b', d):
        return 'E-Sports Props'
    if re.search(r'\baces?\b|' + _RE_FALTAS_TENIS.pattern + r'|break ?points?', d):
        return 'Player Props'                # §6 Tênis: aces / double faults
    if re.search(r'\bstrikeouts?\b', d):
        return 'Player Props'                # §6 Baseball: estatística individual
    if re.search(r'\bruns?\b|\bhits?\b|\bbases\b|\binning', d):
        return 'Corridas'                    # §3: corridas e estatísticas de Baseball
    if re.search(r'\bassi(st|t)\w*', d):
        return 'Assistência' if esporte == 'Futebol' else 'Player Props'
    if re.search(r'\brebotes?\b', d):
        return 'Player Props'
    if re.search(r'cestas|arremessos|\b3p\b|\btres\b', d):
        return 'Team Props'
    if re.search(r'\bfaltas?\b', d):
        return 'Faltas'
    if re.search(r'180s?\b|\bfours?\b', d):
        # §6 Dardos: comparativo ("ter mais 180s") é H2H; total individual é Props.
        return 'H2H' if re.search(r'\bmais\b', d) else 'Player Props'
    if re.search(r'\brounds?\b|\bassaltos?\b', d):
        return 'Rounds'
    if re.search(r'\bpontos?\b|\bpts\b', d):  # ANTES de Sets (§7: unidade contada)
        return 'Pontos'
    if re.search(r'\bsets?\b|\bset1\b|\b1set\b', d):
        return 'Sets'
    # placar exato de sets: "Carlos ganhar por 2-0", "Ya Yi por 2 a 0"
    if esporte in ('Tênis', 'Tênis de Mesa', 'Badminton') and re.search(r'\bpor \d\s*[-a]\s*\d', d):
        return 'Sets'

    # 3. Mercados de RESULTADO (nenhum objeto próprio foi nomeado).
    if _RE_DC.search(d):
        return 'Dupla Chance'
    if re.search(r'\bdnb\b|empate anula', d):
        return 'DNB'
    if re.search(r'\bambas\b', d):
        return 'Ambas Marcam'
    if re.search(r'\bha\b|handicap|spread', d):
        return 'Handicap'
    if re.search(r'\bml\b|moneyline|vencedor|\bempate\b|vence\w*|classific\w*', d):
        return 'ML'

    # 4. Sem objeto nomeado: com linha numérica é total (objeto padrão do
    #    esporte); sem linha, é nome de time solto — resultado.
    if _RE_TOTAL.search(d):
        return _TOTAL_POR_ESPORTE.get(esporte, 'Outros')
    return 'ML'


# ---------- data / stake / odd ----------
def norm_data(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def _para_float(v):
    # bool ANTES de int: célula de fórmula vazia vem como FALSE, e
    # `isinstance(False, int)` é True — sem esta guarda viraria 0,0 silencioso.
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = limpa(v).replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def fmt_stake(v) -> str:
    n = _para_float(v)
    return '' if n is None else f'{n:.2f}'.replace('.', ',')


def _float_str(n: float) -> str:
    """Precisão completa (nunca truncar odd), vírgula decimal."""
    s = repr(float(n))
    if s.endswith('.0'):
        s = s[:-2]
    return s.replace(',', '').replace('.', ',')


def norm_odd(v) -> str:
    n = _para_float(v)
    return '' if n is None or n <= 0 else _float_str(n)


_RESULTADO_MAP = {'green': 'W', 'red': 'L', 'void': 'V',
                  'anulado': 'V', 'anulada': 'V', 'cancelada': 'V'}


def norm_resultado(v) -> str:
    r = _RESULTADO_MAP.get(_chave(v), _chave(v).upper())
    return r if r in VALID else ''


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    faltando = [a for a in ABAS if a not in wb.sheetnames]
    if faltando:
        raise SystemExit(f'aba(s) não encontrada(s): {faltando} — abas: {wb.sheetnames}')

    out: list[dict] = []
    for aba in ABAS:
        ws = wb[aba]
        brutas = [r for r in ws.iter_rows(min_row=LINHA_DADOS, min_col=COL_INI,
                                          max_col=COL_FIM, values_only=True)
                  if r and isinstance(r[0], dt.datetime)]
        # 1 aba = 1 mês é a premissa da numeração mensal. Se cair, a série sai
        # errada — aborta em vez de numerar torto em silêncio.
        meses = {(r[0].year, r[0].month) for r in brutas}
        if len(meses) != 1:
            raise SystemExit(
                f'aba {aba!r} mistura {len(meses)} meses ({sorted(meses)}) — a '
                f'numeração mensal pressupõe 1 aba = 1 mês. Revise antes de importar.')
        # a numeração segue a ORDEM DAS LINHAS; se elas não forem cronológicas a
        # série deixa de refletir a ordem de publicação no canal.
        datas = [r[0] for r in brutas]
        fora = sum(1 for a, b in zip(datas, datas[1:]) if b < a)
        if fora:
            raise SystemExit(
                f'aba {aba!r} tem {fora} linha(s) fora de ordem cronológica — a '
                f'numeração PT<aaaamm>-<n> pressupõe ordem de digitação = ordem '
                f'de publicação. Revise antes de importar.')
        for r in brutas:
            esporte_bruto = limpa(r[2])
            esporte = norm_esporte(esporte_bruto)
            descricao = limpa(r[3])
            out.append({
                'aba': aba,
                'data': norm_data(r[0]),
                '_dt': r[0],
                'esporte': esporte,
                'tipster': TIPSTER,
                'casa': norm_casa(r[1]),
                'parceiro': PARCEIRO,
                'aposta': norm_categoria(esporte_bruto, esporte, descricao),
                'descricao': descricao,
                'stake': fmt_stake(r[6]),
                'odd': norm_odd(r[4]),
                'resultado': norm_resultado(r[5]),
                '_lucro': _para_float(r[7]),      # só para a conferência do DRY
                '_esporte_bruto': esporte_bruto,
                '_casa_bruta': limpa(r[1]),
            })
    return numerar(out)


def numerar(rows: list[dict]) -> list[dict]:
    """Código PT<aaaamm>-<n>. Numeração MENSAL (reinicia em 1 a cada mês), na
    ordem das linhas da planilha dentro do mês."""
    contador: dict[str, int] = defaultdict(int)
    for r in rows:
        mes = f"{r['_dt']:%Y%m}"
        contador[mes] += 1
        r['codigo'] = f'{PREFIXO}{mes}-{contador[mes]}'
        r['_mes'] = mes
    return rows


# ---------- assinatura (idêntica a repository._assinatura) ----------
# COM código o hash é `ID|casa|parceiro|codigo` — o CONTEÚDO não entra. É o que
# faz o bot casar a linha certa quando reprocessar o mesmo número.
def assinatura(r: dict) -> str:
    raw = '|'.join(['ID', r['casa'], r['parceiro'], r['codigo']])
    return hashlib.sha256(raw.encode()).hexdigest()[:20]


def estado_extracao(resultado: str, odd: str) -> str:
    """Espelha repository.estado_extracao: desde a s259 a odd só é exigida onde o
    P/L depende dela (W/HW); L/V/HL nascem `resolvida`."""
    if resultado not in VALID:
        return 'aberta'
    if resultado in ('W', 'HW'):
        return 'resolvida' if (_para_float(odd) or 0) > 0 else 'aberta'
    return 'resolvida'


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict]):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    casas = sorted({r['casa'] for r in rows})

    registros = [(
        DONO, r['casa'], r['parceiro'], assinatura(r), r['codigo'],
        r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
        r['stake'], r['odd'], r['resultado'] or None,
        estado_extracao(r['resultado'], r['odd']),
        None, None, ORIGEM,                       # confianca, stake_usd, origem
    ) for r in rows]

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=120)
            try:
                async with conn.transaction():
                    # idempotente: reimportar não acumula (limpa só o que ESTE
                    # import escreveu — captura do bot/extensão tem outra origem)
                    apagadas = await conn.execute(
                        'DELETE FROM bilhetes WHERE dono=$1 AND origem=$2', DONO, ORIGEM)
                    print(f'  [tentativa {tentativa}] limpou import anterior: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    # uma conta `Padrão` POR CASA: 1 linha por casa no Painel de
                    # Contas, cada uma com custo próprio
                    for casa in casas:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            DONO, casa, PARCEIRO)
                    # feed ordena por criado_em DESC; num import não existe
                    # "envio" → ancora na data da aposta para sair cronológico
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC,
                                                               id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem=$2
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, DONO, ORIGEM)
                n = await conn.fetchval('SELECT COUNT(*) FROM bilhetes WHERE dono=$1', DONO)
                nc = await conn.fetchval(
                    'SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1', DONO)
                np = await conn.fetchval('SELECT COUNT(*) FROM parceiros WHERE dono=$1', DONO)
                print(f'\nOK — bilhetes dono={DONO}={n} | casas={nc} | contas={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


# ---------- relatório do DRY ----------
def _pl_derivado(stake: str, odd: str, res: str):
    s = _para_float(stake) or 0.0
    o = _para_float(odd)
    if res == 'L':
        return -s
    if res == 'V':
        return 0.0
    if res == 'HL':
        return -s / 2
    if o is None:
        return None
    if res == 'W':
        return s * (o - 1)
    if res == 'HW':
        return s * (o - 1) / 2
    return None


def _relatorio(rows: list[dict]):
    print(f'DONO={DONO} | tipster={TIPSTER!r} | conta={PARCEIRO!r} por casa | '
          f'linhas: {len(rows)}')

    print('\n— por aba —')
    for aba in ABAS:
        sub = [r for r in rows if r['aba'] == aba]
        if not sub:
            continue
        datas = sorted(r['_dt'] for r in sub)
        print(f'  {aba:<8} {len(sub):>4} | {datas[0]:%d/%m/%Y} → {datas[-1]:%d/%m/%Y} '
              f'| códigos {sub[0]["codigo"]} … {sub[-1]["codigo"]}')

    for campo in ('casa', 'esporte', 'aposta'):
        print(f'\n{campo}:', dict(Counter(r[campo] for r in rows).most_common()))
    print('\nresultado:',
          dict(Counter(r['resultado'] or '(aberta)' for r in rows).most_common()))

    novas = {r['casa'] for r in rows} - set(_CASA_MAP.values())
    if novas:
        print(f'\n⚠ casa(s) fora do mapa, gravadas VERBATIM: {sorted(novas)}')
    tem_novas = sorted(CASAS_NOVAS & {r['casa'] for r in rows})
    if tem_novas:
        print(f'\n⚠ casa(s) NOVAS no sistema (não existiam nas 59 grafias do banco): '
              f'{tem_novas} — cada uma precisa de favicon nos 3 mapas '
              f'(index.html, data.js, inicio.html).')

    nc = Counter(r['esporte'] for r in rows if r['esporte'] in NAO_CANONICOS)
    if nc:
        print(f'\n⚠ esporte(s) FORA do MASTER_ESPORTES §7 (entram verbatim; a rota '
              f'/taxonomia serve a união com a base do dono): {dict(nc)}')

    sigs = [assinatura(r) for r in rows]
    cods = [r['codigo'] for r in rows]
    print(f'\ncódigos: {len(set(cods))} únicos de {len(cods)}')
    print(f'assinaturas: {len(set(sigs))} únicas de {len(sigs)}')

    print('extraction_state:', dict(Counter(
        estado_extracao(r['resultado'], r['odd']) for r in rows)))

    sem_odd = [r for r in rows if not r['odd']]
    if sem_odd:
        print(f'\n⚠ {len(sem_odd)} linha(s) sem odd')
    sem_stake = [r for r in rows if (_para_float(r['stake']) or 0) <= 0]
    if sem_stake:
        print(f'\n⚠ {len(sem_stake)} linha(s) com stake 0 ou vazio: gravadas, porém '
              f'INVISÍVEIS no dashboard (dashboard_rows corta stake <= 0):')
        for r in sem_stake[:10]:
            print(f'    {r["codigo"]} | {r["data"]} | {r["descricao"][:50]} | '
                  f'stake={r["stake"]!r}')

    # conferência contra a coluna Lucro da própria planilha, linha a linha: se o
    # P/L derivado divergir, a classificação ou a normalização quebrou.
    # `void` grava a string "VOID" na coluna Lucro → _lucro=None e o P/L é 0.
    liq = [r for r in rows if r['resultado']]
    div = [r for r in liq
           if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is None
           or abs(v - (r['_lucro'] or 0)) > 0.02]
    print(f'\nP/L derivado × coluna "Lucro" da planilha: '
          f'{len(div)} divergência(s) em {len(liq)} liquidadas')
    for r in div[:12]:
        print(f'    {r["codigo"]} | {r["data"]} | {r["resultado"]} | '
              f'{r["descricao"][:40]} | u={r["stake"]} @{r["odd"]} | '
              f'planilha={r["_lucro"]} | derivado='
              f'{_pl_derivado(r["stake"], r["odd"], r["resultado"])}')

    turnover = sum(_para_float(r['stake']) or 0 for r in rows)
    to_liq = sum(_para_float(r['stake']) or 0 for r in liq)
    pl = sum(v for r in liq
             if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is not None)
    print(f'\nturnover total (u):      {turnover:>10,.2f}')
    print(f'turnover liquidado (u):  {to_liq:>10,.2f}')
    print(f'P/L derivado (u):        {pl:>10,.2f}   '
          f'(ROI {pl / to_liq * 100:.1f}% sobre o liquidado)')

    # o mesmo P/L por MÊS, contra o "LUCRO DO MÊS" que a planilha estampa no topo
    print('\n— P/L por mês (comparar com o "LUCRO DO MÊS" do topo de cada aba) —')
    for mes in sorted({r['_mes'] for r in rows}):
        sub = [r for r in rows if r['_mes'] == mes and r['resultado']]
        p = sum(v for r in sub
                if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is not None)
        t = sum(_para_float(r['stake']) or 0 for r in sub)
        print(f'  {mes}: {len(sub):>4} liquidadas | turnover {t:>8,.2f}u | '
              f'P/L {p:>8,.2f}u | ROI {p / t * 100:5.2f}%')

    print('\n— rótulo bruto da coluna Esporte → (esporte, e as categorias que gerou) —')
    pares = defaultdict(Counter)
    for r in rows:
        pares[(r['_esporte_bruto'], r['esporte'])][r['aposta']] += 1
    for (bruto, esp), cats in sorted(pares.items(), key=lambda kv: -sum(kv[1].values())):
        marca = ''
        if esp in NAO_CANONICOS:
            marca = '  ⚠ fora do MASTER §7'
        elif _k(bruto) in PSEUDO_ESPORTES:
            marca = '  ← não é esporte, é tipo de bilhete'
        n = sum(cats.values())
        print(f'  {bruto:<18} → {esp:<18} ({n:>3}) {dict(cats.most_common(6))}{marca}')

    outros = [r for r in rows if r['aposta'] == 'Outros']
    print(f'\n— categoria `Outros`: {len(outros)} linha(s) '
          f'({len(outros) / len(rows) * 100:.1f}%) —')
    for r in outros:
        print(f'    {r["codigo"]} | {r["esporte"]:<12} | {r["descricao"]}')

    print('\n=== 16 AMOSTRAS (Código|Data|Esporte|Casa|Aposta|Descrição|Stake|Odd|Res) ===')
    step = max(1, len(rows) // 16)
    for r in rows[::step][:16]:
        print(' | '.join([r['codigo'], r['data'], r['esporte'], r['casa'], r['aposta'],
                          r['descricao'][:42], r['stake'], r['odd'],
                          r['resultado'] or '—']))

    ultimo = max((r for r in rows if r['_mes'] == '202608'),
                 key=lambda r: int(r['codigo'].rsplit('-', 1)[1]), default=None)
    if ultimo:
        n = int(ultimo['codigo'].rsplit('-', 1)[1])
        print(f'\n>>> CONTADOR DO BOT — data/passatips/contador.json: '
              f'{{"2026-08": {n}}}  (próxima do canal = #{n + 1})')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows))


if __name__ == '__main__':
    main()
