# -*- coding: utf-8 -*-
"""Importa a base PESSOAL do Lava (.xlsx, aba 'BASE') para dono='LavaPessoal'.

NÃO confundir com `import_lava.py`: aquele é o operador `Lava` (base do CSV
compartilhado, sessão 75). Este é o perfil **pessoal e isolado** `LavaPessoal`
(dono solo desde a s216 — ver `app/auth.py`). São duas bases distintas que só
compartilham o apelido.

Layout da planilha (header na linha 3, dados começam na 4):

    0 Data | 1 Esporte | 2 Tipster | 3 CASA | 4 Aposta | 5 Descrição |
    6 Stake (R$) | 7 Odd | 8 W/L | 9 U Investida | 10 P/L (U) | 11 P/L (R$)

Não existe coluna Parceiro — o Lava não anota fornecedor. Decisão do Feca:
toda conta entra como `Padrão` (sem sufixo `[Fornecedor]`), uma por casa.

As colunas 9/10/11 são DESCARTADAS. A planilha contabiliza em "unidade", e o
valor em R$ da unidade não acompanha a coluna `Stake (R$)`: em 2553 das 2860
linhas liquidadas o `P/L (R$)` não é `stake × (odd−1)` (ex.: stake 800, odd
1,83, W → a planilha marca R$ 830; o derivado é R$ 664). O P/L do sistema é
DERIVADO (`calcular_pl`, nunca persistido), então o dashboard mostra o número
de stake × odd. Não há como conciliar sem falsificar a stake.

Uso:
    python scripts/import_lavapessoal_xlsx.py --xlsx "C:\\...\\lava.xlsx"        # DRY
    python scripts/import_lavapessoal_xlsx.py --xlsx "C:\\...\\lava.xlsx" --go   # escreve
"""
import argparse
import asyncio
import datetime as dt
import hashlib
import os
import re
from collections import Counter

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
DONO = 'LavaPessoal'
PARCEIRO = 'Padrão'          # sem fornecedor (decisão do Feca)
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}


# ---------- sanitização de texto ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')       # C0 + DEL (inclui \n \r \t \x0c)
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')      # escape literal do openpyxl (_x000C_)


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


# ---------- casa ----------
# `casa` é TEXTO em 7 tabelas: cada grafia é uma casa DIFERENTE no sistema
# (CLAUDE.md). Por isso o de-para é explícito e casa desconhecida entra VERBATIM
# — title-casear nome novo cria conta paralela.
_CASA_MAP = {
    '365': 'Bet365', '365.0': 'Bet365', 'BET365': 'Bet365',
    'BETANO': 'Betano',
    'SUPERBET': 'Superbet',
    'BETFAST': 'Betfast',
    'BETNACIONAL': 'Betnacional',
    'KTO': 'KTO',
    'SPORTINGBET': 'SportingBet',
    'PINNACLE': 'Pinnacle',
    'BETFAIR': 'Betfair',
    'BINGOPLUS': 'BingoPlus',
    'KINGPANDA': 'KingPanda',
    # decisões do Feca nesta base
    'TIVOBET': 'Tivo',           # casa canônica do projeto (CASA_TIVO.md)
    'REI': 'Pitaco',             # "Rei do Pitaco" virou "Pitaco"
    # casas próprias desta base (verbatim, sem canon prévio no sistema)
    'FAZ1BET': 'Faz1bet', 'APOSTA1': 'Aposta1', 'BETFUSION': 'Betfusion',
    'STAKE': 'Stake', 'MAXIMABET': 'Maximabet', 'MCGAMES': 'Mcgames',
}


def norm_casa(c) -> str:
    bruto = limpa(c)
    return _CASA_MAP.get(bruto.upper().replace(' ', ''), bruto)


# ---------- esporte ----------
_ESPORTE_MAP = {
    'futebol': 'Futebol',
    'misto': 'Múltiplos',            # mistura de modalidades (MASTER_ESPORTES §2)
    'basquete': 'Basquete',
    'volei': 'Vôlei', 'vôlei': 'Vôlei',
    'formula 1': 'Fórmula 1', 'fórmula 1': 'Fórmula 1',
    'tenis': 'Tênis', 'tênis': 'Tênis',
    'handebol': 'Handebol',
    # 1 linha veio com o cabeçalho repetido na célula; é do tipster Cantos,
    # que só aposta em futebol (conferido: 527 linhas, todas Futebol).
    'esporte': 'Futebol',
}


def norm_esporte(esp) -> str:
    e = limpa(esp)
    return _ESPORTE_MAP.get(e.lower(), e)


# ---------- tipster ----------
# A planilha grava tudo em CAIXA ALTA. Decisão do Feca: Title Case no dashboard.
_TIPSTER_MAP = {
    'arrudex': 'Arrudex', 'peixe': 'Peixe', 'cantos': 'Cantos',
    'minisilva': 'MiniSilva', 'iranian': 'Iranian', 'marco': 'Marco',
    'latino': 'Latino', 'araujo': 'Araujo',
}


def norm_tipster(t) -> str:
    b = limpa(t)
    return _TIPSTER_MAP.get(b.lower(), b.title())


# ---------- categoria (coluna `aposta`) ----------
# A descrição desta base é quase sempre só o nome do time ("Fulham", "City"), então
# o motor de keywords do `import_dashboard_xlsx` não serve: o objeto da aposta vem do
# RÓTULO + do TIPSTER, que aqui é especializado por mercado (decisão do Feca).
_MULTI = {'múltipla', 'multipla', 'aumentada'}

# Cantos = tipster de escanteios (529 linhas). Estes rótulos dele são todos o mesmo
# objeto — escanteio — variando só o tipo de mercado, que pelo MASTER_APOSTAS §1 não
# altera a categoria ("Time B mais escanteios" (comparativo) → Escanteios).
#   Over/Under = total de escanteios · Race = corrida até o N-ésimo escanteio
#   HTML / MLHT = escanteios do 1º tempo · Last = último escanteio
_CANTOS_ESCANTEIOS = {'over', 'under', 'race', 'html', 'mlht', 'handicap', 'last'}


def categoria(tipster: str, esporte: str, rotulo, descricao: str) -> str:
    rl = limpa(rotulo).lower()
    d = descricao.lower()

    # 1) cupom combinado — o rótulo manda, e ` // ` é o separador de seleção (#19)
    if rl in _MULTI or ' // ' in f' {descricao} ':
        return 'Múltipla'

    # 2) tipster especializado em escanteios
    if tipster == 'Cantos' and rl in _CANTOS_ESCANTEIOS:
        return 'Escanteios'

    # 3) Fórmula 1 (tipster Marco). ATENÇÃO ao rótulo "Pontos" aqui: é "terminar nos
    #    pontos" (posição do piloto), NÃO a categoria `Pontos`, que é o objeto ponto
    #    de Basquete/Vôlei/eBasket/Badminton (MASTER_APOSTAS §5).
    if esporte == 'Fórmula 1':
        if rl == 'comparativo':                 # "Norris por delante de Leclerc"
            return 'H2H'
        if rl in ('corrida', 'ml'):             # vencedor / líder / carro vencedor
            return 'ML'
        return 'Player Props'                   # Classificação, Pódio, Top2/3, Posição…

    # 4) rótulos que já nomeiam a categoria
    if rl in ('sot',):
        return 'Chutes no Gol'
    if rl == 'cartão':
        return 'Cartões'
    if rl == 'prop player':
        return 'Player Props'
    if rl in ('handicap', 'handcap'):
        # "Handicap (Sets) Brazil (W) +1.5" — no Vôlei a unidade contada decide
        # entre Sets e Pontos (MASTER_APOSTAS §5, Discriminante 2).
        return 'Sets' if 'set' in d else 'Handicap'
    if rl == 'ml':
        return 'ML'
    if rl == 'anytime':
        return 'Anytime'

    # 5) total (over/under) — a categoria é o objeto medido, que vem do esporte
    if rl in ('over', 'under'):
        if esporte in ('Basquete', 'Vôlei'):
            return 'Pontos'                     # total do jogo ou do time (§5)
        if esporte == 'Futebol':
            return 'Gols'
        return 'Outros'

    return 'Outros'


# ---------- data / stake / odd ----------
def norm_data(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def _para_float(v):
    """Número da planilha → float. Absorve a odd gravada como moeda ("R$ 2,25",
    9 linhas de 21/07): o valor é a odd, só o formato da célula está errado."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = limpa(v).replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def fmt_stake(v) -> str:
    n = _para_float(v)
    return '' if n is None else f'{n:.2f}'.replace('.', ',')


def fmt_odd(v) -> str:
    """Odd com a precisão completa que veio da planilha (nunca truncar — o valor
    entra no P/L derivado). `repr` do float dá a representação mais curta que faz
    roundtrip, então 1.83 sai "1,83" e 43.07 sai "43,07"."""
    n = _para_float(v)
    if n is None:
        return ''
    s = repr(n)
    if s.endswith('.0'):
        s = s[:-2]
    return s.replace('.', ',')


def norm_resultado(v) -> str:
    r = limpa(v).upper()
    return r if r in VALID else ''


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if 'BASE' not in wb.sheetnames:
        raise SystemExit(f"aba 'BASE' não encontrada — abas: {wb.sheetnames}")
    ws = wb['BASE']
    # A linha de dado é a que tem data REAL na coluna 0. Isso descarta de uma vez o
    # bilhete de recado do topo ("Peixe | Preciso renovar meu app…"), o cabeçalho e
    # as ~34 mil linhas vazias que o Sheets exporta.
    brutas = [r for r in ws.iter_rows(values_only=True)
              if r and isinstance(r[0], dt.datetime)]
    out = []
    for r in brutas:
        esporte = norm_esporte(r[1])
        tipster = norm_tipster(r[2])
        descricao = limpa(r[5])
        out.append({
            'data': norm_data(r[0]),
            'esporte': esporte,
            'tipster': tipster,
            'casa': norm_casa(r[3]),
            'parceiro': PARCEIRO,
            'aposta': categoria(tipster, esporte, r[4], descricao),
            'descricao': descricao,
            'stake': fmt_stake(r[6]),
            'odd': fmt_odd(r[7]),
            'resultado': norm_resultado(r[8]),
        })
    return out


# ---------- assinatura (idêntica a repository._assinatura / _norm_odd) ----------
# ATENÇÃO: `stake` ENTRA no hash. Os importadores antigos (`import_lava.py`,
# `import_dashboard_xlsx.py`) copiaram a versão pré-s133, sem stake — não usar
# aqueles como modelo. Divergir daqui faz a próxima captura gerar assinatura nova,
# não colidir, e duplicar o histórico inteiro.
def _norm_odd(v: str) -> str:
    try:
        return f"{round(float(v.replace(',', '.')), 2):.2f}"
    except (ValueError, AttributeError):
        return v


def assinaturas(rows: list[dict]) -> list[str]:
    """Assinatura por linha com contador de duplicatas por lote (igual ao app):
    conteúdo idêntico ESCALA o contador, nunca colide (senão um sobrescreve o outro
    em silêncio)."""
    counts: dict[str, int] = {}
    sigs = []
    for r in rows:
        base_raw = "|".join([
            r['casa'], r['parceiro'], r['data'], r['aposta'], r['descricao'],
            r['stake'], _norm_odd(r['odd']),
        ])
        base_sig = hashlib.sha256(base_raw.encode()).hexdigest()[:20]
        cnt = counts.get(base_sig, 0) + 1
        counts[base_sig] = cnt
        raw = base_raw if cnt == 1 else f"{base_raw}|{cnt}"
        sigs.append(hashlib.sha256(raw.encode()).hexdigest()[:20])
    return sigs


def estado_extracao(resultado: str, odd: str) -> str:
    """Espelha repository.estado_extracao: 'resolvida' EXIGE resultado canônico E
    odd > 0. Sem odd, a linha fica 'aberta' — não polui o feed nem duplica calada."""
    if resultado not in VALID:
        return 'aberta'
    n = _para_float(odd)
    return 'resolvida' if (n or 0) > 0 else 'aberta'


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict]):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    sigs = assinaturas(rows)
    casas = sorted({r['casa'] for r in rows})

    registros = []
    for r, sig in zip(rows, sigs):
        registros.append((
            DONO, r['casa'], r['parceiro'], sig, None,           # codigo_bilhete
            r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
            r['stake'], r['odd'], r['resultado'] or None,
            estado_extracao(r['resultado'], r['odd']),
            None, None, ORIGEM,                                  # confianca, stake_usd
        ))

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=120)
            try:
                async with conn.transaction():
                    # idempotente: reimportar não acumula (limpa só o que ESTE import
                    # escreveu — captura da extensão tem origem='extracao' e sobrevive)
                    apagadas = await conn.execute(
                        "DELETE FROM bilhetes WHERE dono=$1 AND origem=$2", DONO, ORIGEM)
                    print(f'  [tentativa {tentativa}] limpou import anterior: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    for casa in casas:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            DONO, casa, PARCEIRO)
                    # o feed ordena por criado_em DESC (ordem de ENVIO). Num import não
                    # existe "envio", então ancora na data da aposta para a lista sair
                    # cronológica em vez de embaralhada.
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC,
                                                               id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem=$2
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, DONO, ORIGEM)
                n = await conn.fetchval("SELECT COUNT(*) FROM bilhetes WHERE dono=$1", DONO)
                nc = await conn.fetchval(
                    "SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1", DONO)
                np = await conn.fetchval("SELECT COUNT(*) FROM parceiros WHERE dono=$1", DONO)
                print(f'\nOK — bilhetes dono={DONO}={n} | casas={nc} | contas={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


# ---------- relatório do DRY ----------
def _pl_derivado(stake: str, odd: str, res: str):
    s = _para_float(stake) or 0.0
    o = _para_float(odd)
    if res == 'L':
        return -s
    if res == 'V':
        return 0.0
    if o is None or o <= 0:
        return None
    if res == 'W':
        return s * (o - 1)
    if res == 'HW':
        return s * (o - 1) / 2
    if res == 'HL':
        return -s / 2
    return None


def _relatorio(rows: list[dict]):
    print(f'DONO={DONO} | conta={PARCEIRO!r} | linhas: {len(rows)}')
    for campo in ('casa', 'esporte', 'tipster', 'aposta'):
        print(f'\n{campo}:', dict(Counter(r[campo] for r in rows).most_common()))
    print('\nresultado:',
          dict(Counter(r['resultado'] or '(aberta)' for r in rows).most_common()))

    sigs = assinaturas(rows)
    dup = len(sigs) - len(set(sigs))
    print(f'\nassinaturas: {len(set(sigs))} únicas de {len(sigs)} '
          f'({dup} desambiguadas por contador)')

    estados = Counter(estado_extracao(r['resultado'], r['odd']) for r in rows)
    print('extraction_state:', dict(estados))

    # dashboard_rows descarta stake <= 0 — linha assim fica no banco e some do feed
    sem_stake = [r for r in rows if (_para_float(r['stake']) or 0) <= 0]
    if sem_stake:
        print(f'\n⚠ {len(sem_stake)} linha(s) com stake 0: gravadas, porém INVISÍVEIS '
              f'no dashboard (dashboard_rows corta stake <= 0).')
        for r in sem_stake:
            print(f'    {r["data"]} | {r["tipster"]} | {r["casa"]} | {r["descricao"][:52]}')

    turnover = sum(_para_float(r['stake']) or 0 for r in rows)
    # `is not None` e não truthiness: V devolve 0.0, que é falsy — filtrar por verdade
    # descartaria a linha em silêncio (aqui é neutro na soma, mas a forma errada vira
    # bug quando alguém copiar este laço para contar linhas).
    pl = sum(v for r in rows
             if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is not None)
    print(f'\nturnover:      R$ {turnover:>14,.2f}')
    print(f'P/L derivado:  R$ {pl:>14,.2f}   (o que o dashboard vai mostrar)')

    print('\n=== 14 AMOSTRAS (Data|Esporte|Tipster|Casa|Conta|Aposta|Descrição|Stake|Odd|Res) ===')
    step = max(1, len(rows) // 14)
    for r in rows[::step][:14]:
        print(' | '.join([r['data'], r['esporte'], r['tipster'], r['casa'], r['parceiro'],
                          r['aposta'], r['descricao'][:40], r['stake'], r['odd'],
                          r['resultado'] or '—']))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows))


if __name__ == '__main__':
    main()
