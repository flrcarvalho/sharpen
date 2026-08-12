# -*- coding: utf-8 -*-
"""Importa a base do tipster Rei do Criquete (REI DO CRIQUETE PLANILHA FREE.xlsx)
para dono='reidocriquete'.

Marca e username COINCIDEM aqui (`Rei do Criquete` / `reidocriquete`) — conferido
na tabela `usuarios` antes do import, não deduzido do nome do arquivo. É a regra
da s260: o `dono` tem de ser o username, senão a base fica invisível para ele e o
sintoma não é erro, é tela vazia.

A conta foi criada pelo PRÓPRIO usuário no site (Fase 2 do onboarding) e aprovada
pelo Feca — `status='ativo'`, hash de 60 caracteres que ele mesmo escolheu.
**Não existe env var nem linha em `app/auth.py` para esta conta.**

── Layout da planilha ────────────────────────────────────────────────────────

Três abas com apostas (`KING March`, `KING April`, `KING Agosto`). Em todas, a
linha 44 é o cabeçalho e os dados começam na 46. Só as colunas A..K são apostas;
da M em diante são painéis de KPI da própria planilha (tabelas auxiliares de
"Ganhos por dia / por mercado / por unidade / por casa") e são IGNORADAS:

    0 Data | 1 Ao-vivo/Pre | 2 Competição | 3 Time 1 | 4 Time 2 | 5 Mercado |
    6 Unidades | 7 ODDS | 8 Resultado | 9 Lucro Obtido (Un) | 10 Casa de Aposta

A linha de dado é reconhecida por `Data` REAL na coluna 0 — o mesmo critério do
`import_fleury_xlsx.py`, que descarta cabeçalho, linhas vazias e os painéis.

── O que a planilha NÃO registra (limitação da fonte, não do import) ─────────

A coluna `Mercado` é uma FAMÍLIA grosseira (`Over`/`Under`/`ML`/`Sim`/`Não`/`X`/
`Handicap`/`Anytime`/`Card`/`MULT`), sem a linha e sem o lado escolhido. Três
linhas do mesmo jogo (`LA Lakers v HOU Rockets`, mercado `Não`) trazem odds
**3,40 · 19,00 · 81,00** — o rótulo não identifica a aposta. Idem `ML`, que não
diz qual time.

Por isso a descrição sai como `<Mercado> [Time 1 v Time 2]` — o máximo honesto
sobre o que a fonte tem. O `MASTER_DESCRICAO §2` pede
`Entidade - Mercado [Confronto]`; aqui não existe Entidade nem linha, e o §1 é
explícito: **nunca inventar informações**. Preencher isso exigiria os prints
originais do grupo, que não temos.

── Esporte ───────────────────────────────────────────────────────────────────

A coluna `Competição` mistura ESPORTE (`Futebol`, `Basquete`, `Tenis`…) com
MERCADO (`CANTOS`, `Props`, `MULTIPLA`). Os três últimos são de futebol/multi:

- `CANTOS`/`Cantos` → escanteios, esporte `Futebol`
- `Props`          → props de jogador; as 31 linhas são confrontos de futebol
                     (Bragantino v Atlético MG, PSG v Aston Villa…) — MEDIDO,
                     não suposto
- `MULTIPLA`       → esporte `Múltiplos` (as 85 linhas trazem `3 ACIMA` nos dois
                     times, ou seja 3+ seleções; o conteúdo não é registrado)

⚠️ QUATRO esportes NÃO são canônicos no `MASTER_ESPORTES §7` (a lista tem 21 e
nenhum deles está lá): `Críquete` (29 linhas), `Sinuca`, `Polo Aquático` e
`Tênis de Mesa` (1 cada). Eles entram VERBATIM, que é o precedente documentado
no `CLAUDE.md` — a base preserva a grafia herdada de import (`Fórmula 1`,
`Esoccer`, `Tênis de Mesa`) e a rota `/taxonomia` serve a UNIÃO do canônico com
a base do dono, então os menus dele mostram os quatro. As grafias escolhidas são
as que o sistema JÁ conhece: `Críquete` está no mapa de favicon do `index.html`
e `Tênis de Mesa` é citado no `CLAUDE.md`.

**Não** mexi no MASTER: criar esporte é a regra de propagação (mudança própria,
com aprovação). Fica anotado — o tipster se chama Rei do Criquete e Críquete não
existe na taxonomia canônica.

── Categoria ─────────────────────────────────────────────────────────────────

Derivada do par (Competição, Mercado). Onde o MASTER tem correspondência
inequívoca, usa-se ela; onde não tem, `Outros` — nunca um palpite:

- `MULT`                → `Múltipla`
- `ML`                  → `ML`
- `X`/`x`               → `ML` (o empate do 1X2; aparece em Beisebol japonês/
                          coreano e Handebol, onde empate existe, com odds
                          21–29 e 13–14 — compatível)
- `Handicap`            → `Handicap`
- `Card`                → `Cartões` (§: cartão de jogador é `Cartões`, NUNCA
                          `Player Props` — o objeto apostado é o cartão)
- `Anytime`             → `Anytime`
- `Assistencia`         → `Assistência`
- `Sim`/`Não`/`Nao`     → `Outros` ⚠️ 78 linhas. NÃO é Ambas Marcam: as odds
                          (3,4 · 19 · 81 no mesmo jogo de basquete; 17 · 19 em
                          futebol) descartam. É atalho do tipster para algo que
                          a planilha não registra. Perguntar ao Fred/Kauan.
- `Over`/`Under`        → depende do esporte (o objeto apostado muda):
      CANTOS → `Escanteios` · Props → `Player Props` · Futebol → `Gols` ·
      Basquete/Beisebol → `Pontos` · UFC → `Rounds` · E-Sports →
      `E-Sports Props` · Tênis → `Games` · resto → `Outros`

── Numeração (decisão do Feca) ───────────────────────────────────────────────

Código `RC<aaaamm>-<n>`, na família do `SC` (Só Chutes) e `ZE` (Zora eSports).
A numeração é MENSAL: reinicia em 1 a cada mês e fecha no fim do mês. O `n` segue
a ORDEM DAS LINHAS da planilha dentro do mês — a coluna Data é a do EVENTO e está
fora de ordem (12→17/08 embaralhados), então a ordem de digitação é a melhor
proxy da ordem de publicação no grupo.

O mês sai da DATA da linha, não do nome da aba, e o script ABORTA se uma aba
misturar meses (aí a premissa "1 aba = 1 mês" caiu e a numeração estaria errada).

Isso não é cosmético: `repository._assinatura` com código é `ID|casa|parceiro|
codigo` — o CONTEÚDO não entra no hash. Numerar é o que faz o bot, ao planilhar
a #66 em diante, casar com a linha certa em vez de duplicar. De quebra resolve as
linhas de conteúdo idêntico (`Platense v Coquimbo · Card · 0,25u @4,50` aparece
duas vezes em agosto): códigos distintos = linhas distintas, sem `_counter`.

── Decisões ──────────────────────────────────────────────────────────────────

- **Dono solo** — sem OPERADORES, sem dedup cruzada.
- **Tipster** = `Rei do Criquete` em todas as linhas (nome de marca).
- **Stake em UNIDADES** (a coluna se chama `Unidades`, 0,10–6,00, e o lucro é
  `Lucro Obtido (Un)`). Importa 1u = 1 — o P/L do dashboard é o P/L em unidades,
  mesma decisão do `SoChutes` e do `Fleury`.
- **Uma conta `Padrão` por casa** — 5 linhas no Painel de Contas, cada uma com
  custo próprio. As 5 grafias já são as canônicas do sistema (conferido no banco:
  Bet365 48.299 · Betano 11.067 · Superbet 5.543 · Betfair 2.155 · Novibet 583).

Resultado: `Certo`/`certo` → W · `errado`/`ERRADO` → L · `anulado`/`Anulado` → V ·
vazio → aberta (54 linhas de agosto, eventos de 12→17/08 ainda por liquidar).

Uso:
    python scripts/import_reidocriquete_xlsx.py --xlsx "C:\\...\\REI DO CRIQUETE PLANILHA FREE.xlsx"        # DRY
    python scripts/import_reidocriquete_xlsx.py --xlsx "C:\\...\\REI DO CRIQUETE PLANILHA FREE.xlsx" --go   # escreve
"""
import argparse
import asyncio
import datetime as dt
import hashlib
import os
import re
import unicodedata
from collections import Counter, defaultdict

import openpyxl

ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env')
DONO = 'reidocriquete'       # username conferido em `usuarios` (marca: Rei do Criquete)
PARCEIRO = 'Padrão'
TIPSTER = 'Rei do Criquete'
PREFIXO = 'RC'               # código RC<aaaamm>-<n>, família do SC/ZE
ABAS = ('KING March', 'KING April', 'KING Agosto')
ORIGEM = 'import'
VALID = {'W', 'L', 'V', 'HW', 'HL'}
LINHA_DADOS = 46             # cabeçalho na 44; da 46 em diante são apostas


# ---------- sanitização de texto ----------
_CTRL = re.compile(r'[\x00-\x1f\x7f]+')
_XESC = re.compile(r'_x[0-9A-Fa-f]{4}_')


def limpa(v) -> str:
    if v is None:
        return ''
    s = _XESC.sub(' ', str(v))
    return re.sub(r'\s{2,}', ' ', _CTRL.sub(' ', s)).strip()


def _chave(s: str) -> str:
    """minúscula sem acento e sem espaço — para casar rótulo digitado à mão
    (`Nao`/`Não`, `MUlt`/`MULT`, `beisebol`/`Beisebol`)."""
    s = unicodedata.normalize('NFKD', limpa(s))
    return ''.join(c for c in s if not unicodedata.combining(c)).lower().replace(' ', '')


# ---------- casa ----------
# `casa` é TEXTO em 7 tabelas: cada grafia é uma casa DIFERENTE (CLAUDE.md).
# As 5 grafias abaixo são as que o banco JÁ usa — conferidas, não supostas.
# Casa desconhecida entra VERBATIM (nunca title-casear: mutila nome e cria conta
# paralela — `BETesporte`, `VaideBet`, `KingPanda`).
_CASA_MAP = {
    'bet365': 'Bet365',
    'novibet': 'Novibet',
    'superbet': 'Superbet',
    'betano': 'Betano',
    'betfair': 'Betfair',
}


def norm_casa(v) -> str:
    bruto = limpa(v)
    return _CASA_MAP.get(_chave(bruto), bruto)


# ---------- esporte ----------
# `Competição` mistura esporte e mercado. CANTOS/Props são futebol; MULTIPLA é
# `Múltiplos`. Os 4 marcados NÃO-CANÔNICO não existem no MASTER_ESPORTES §7 e
# entram verbatim, na grafia que o sistema já conhece.
_ESPORTE_MAP = {
    'futebol': 'Futebol',
    'cantos': 'Futebol',            # escanteios
    'props': 'Futebol',             # 31/31 são confrontos de futebol (medido)
    'multipla': 'Múltiplos',        # `3 ACIMA` nos dois times = 3+ seleções
    'basquete': 'Basquete',
    'esports': 'E-Sports',
    'tenis': 'Tênis',
    'beisebol': 'Baseball',
    'hoquei': 'Hóquei',
    'dardos': 'Dardos',
    'rugby': 'Rugby',
    'voleibol': 'Vôlei',
    'handebol': 'Handebol',
    'ufc': 'MMA',
    'criquete': 'Críquete',         # NÃO-CANÔNICO (grafia do mapa de favicon)
    'sinuca': 'Sinuca',             # NÃO-CANÔNICO
    'poloaquatico': 'Polo Aquático',  # NÃO-CANÔNICO
    'pingpong': 'Tênis de Mesa',    # NÃO-CANÔNICO (grafia citada no CLAUDE.md)
}
NAO_CANONICOS = {'Críquete', 'Sinuca', 'Polo Aquático', 'Tênis de Mesa'}


def norm_esporte(competicao) -> str:
    return _ESPORTE_MAP.get(_chave(competicao), limpa(competicao) or 'Outro')


# ---------- categoria ----------
# Rótulos que não dependem do esporte.
_CAT_DIRETA = {
    'mult': 'Múltipla',
    'ml': 'ML',
    'x': 'ML',                  # empate do 1X2
    'handicap': 'Handicap',
    'card': 'Cartões',          # objeto apostado é o cartão, nunca Player Props
    'anytime': 'Anytime',
    'assistencia': 'Assistência',
    # ⚠️ atalho do tipster que a planilha não explica — ver docstring
    'sim': 'Outros',
    'nao': 'Outros',
}

# Over/Under: a categoria segue o OBJETO, que muda com o esporte (MASTER §1).
_CAT_OVER_POR_COMPETICAO = {
    'cantos': 'Escanteios',
    'props': 'Player Props',
    'futebol': 'Gols',
    'basquete': 'Pontos',
    'beisebol': 'Pontos',
    'ufc': 'Rounds',
    'esports': 'E-Sports Props',
    'tenis': 'Games',
}


def norm_categoria(competicao, mercado) -> str:
    m = _chave(mercado)
    if m in _CAT_DIRETA:
        return _CAT_DIRETA[m]
    if m in ('over', 'under'):
        return _CAT_OVER_POR_COMPETICAO.get(_chave(competicao), 'Outros')
    return 'Outros'


# ---------- descrição ----------
# `<Mercado> [Time 1 v Time 2]`. Sem entidade e sem linha porque a fonte não tem
# (MASTER_DESCRICAO §1: nunca inventar). Rótulo do mercado normalizado só na
# caixa; `X` vira `Empate`, que é o que ele significa.
_MERCADO_ROTULO = {
    'over': 'Over', 'under': 'Under', 'ml': 'ML', 'x': 'Empate',
    'handicap': 'Handicap', 'card': 'Cartão', 'anytime': 'Anytime',
    'assistencia': 'Assistência', 'sim': 'Sim', 'nao': 'Não', 'mult': 'Múltipla',
}

# `Time 2` nem sempre é adversário: em 139 linhas ele é RÓTULO DE AGRUPAMENTO,
# ligando as pernas ao combo. Em 06/04, por exemplo, convivem três linhas
# `Futebol | Idrissa Gueye | TRIPLA 1 | Anytime` (as pernas) e a linha
# `MULTIPLA | TRIPLA 1 | TRIPLA 1 | MULT @318,50` (o combo). Tratar isso como
# confronto produziria `Anytime [Mattia Liberali v TRIPLA 1]` — um adversário
# que não existe, e o `MASTER_DESCRICAO §1` proíbe criar confronto inexistente.
#
# Medido: 114 linhas-combo (marcador dos DOIS lados) e 24 linhas-perna (jogador
# real de um lado, marcador do outro). O `a` solto é lixo de digitação em 6
# linhas de `MULTIPLA` e entra como marcador.
_MARCADOR_COMBO = re.compile(
    r'^(?:\d+\s*ACIMA'
    r'|\d*\s*(?:TRIPLAS?|DUPLAS?|QUADRUPLAS?)(?:\s*\d+)?'
    r'|LINHA\s*\d+'
    r'|a)$', re.I)
_N_ACIMA = re.compile(r'^(\d+)\s*ACIMA$', re.I)


def _rotulo_combo(marcador: str) -> str:
    """`3 ACIMA` → `3+ seleções`; `TRIPLA 1`/`Triplas` ficam verbatim."""
    m = _N_ACIMA.match(marcador)
    return f'{m.group(1)}+ seleções' if m else marcador


def norm_descricao(competicao, mercado, t1, t2) -> str:
    rot = _MERCADO_ROTULO.get(_chave(mercado), limpa(mercado))
    a, b = limpa(t1), limpa(t2)
    comp = _chave(competicao)

    if _MARCADOR_COMBO.match(b):
        marcador = _rotulo_combo(b if not _MARCADOR_COMBO.match(a) or _N_ACIMA.match(b)
                                 else a)
        if _MARCADOR_COMBO.match(a):
            # linha do COMBO. `CANTOS` ganha prefixo porque a informação
            # "escanteios" se perderia: o esporte vira `Futebol` e a categoria
            # `Múltipla`, e nenhum dos dois guarda o objeto apostado.
            # O `a` solto não é marcador de nada — vira parêntese vazio de
            # sentido, então some.
            base = 'Múltipla' if marcador.lower() == 'a' else f'Múltipla ({marcador})'
            return f'Escanteios - {base}' if comp == 'cantos' else base
        # linha de PERNA: entidade real + mercado, com o grupo entre parênteses
        return f'{a} - {rot} ({marcador})'

    # Daqui para baixo os dois lados são conteúdo real — inclusive em `MULTIPLA`,
    # onde 2 linhas trazem confronto/pernas de verdade (`EC Graz 99ers v El
    # Karma`, `Menos de 8,5 v Menos de 19,5`). Tratá-las como rótulo de grupo
    # descartava metade da informação.
    if not a and not b:
        return rot
    if not a or not b:
        return f'{rot} [{a or b}]'
    # Mesmo texto dos dois lados não é confronto — é rótulo repetido para
    # preencher a coluna (`VIX DIVERSOS v VIX DIVERSOS`). Ninguém joga contra
    # si mesmo, e `[X v X]` leria como um jogo que não existe.
    if _chave(a) == _chave(b):
        return f'{rot} ({a})'
    return f'{rot} [{a} v {b}]'


# ---------- data / stake / odd ----------
def norm_data(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime('%d/%m/%Y')
    return limpa(v).split(' ')[0]


def _para_float(v):
    # bool ANTES de int: as células de Lucro vazias vêm como FALSE (fórmula), e
    # `isinstance(False, int)` é True — sem esta guarda viravam 0,0 silenciosos.
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = limpa(v).replace('R$', '').replace('\xa0', '').replace(' ', '')
    if not s:
        return None
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def fmt_stake(v) -> str:
    n = _para_float(v)
    return '' if n is None else f'{n:.2f}'.replace('.', ',')


def _float_str(n: float) -> str:
    """Precisão completa (nunca truncar odd), vírgula decimal."""
    s = repr(float(n))
    if s.endswith('.0'):
        s = s[:-2]
    return s.replace(',', '').replace('.', ',')


def norm_odd(v) -> str:
    n = _para_float(v)
    return '' if n is None or n <= 0 else _float_str(n)


_RESULTADO_MAP = {'certo': 'W', 'errado': 'L', 'anulado': 'V',
                  'void': 'V', 'anulada': 'V', 'cancelada': 'V'}


def norm_resultado(v) -> str:
    r = _RESULTADO_MAP.get(_chave(v), _chave(v).upper())
    return r if r in VALID else ''


# ---------- carga da planilha ----------
def carregar_rows(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    faltando = [a for a in ABAS if a not in wb.sheetnames]
    if faltando:
        raise SystemExit(f'aba(s) não encontrada(s): {faltando} — abas: {wb.sheetnames}')

    out: list[dict] = []
    for aba in ABAS:
        ws = wb[aba]
        brutas = [r for r in ws.iter_rows(min_row=LINHA_DADOS, max_col=11, values_only=True)
                  if r and isinstance(r[0], dt.datetime)]
        # 1 aba = 1 mês é a premissa da numeração mensal. Se cair, a série sai
        # errada — aborta em vez de numerar torto em silêncio.
        meses = {(r[0].year, r[0].month) for r in brutas}
        if len(meses) != 1:
            raise SystemExit(
                f'aba {aba!r} mistura {len(meses)} meses ({sorted(meses)}) — a '
                f'numeração mensal pressupõe 1 aba = 1 mês. Revise antes de importar.')
        for r in brutas:
            competicao, mercado = r[2], r[5]
            out.append({
                'aba': aba,
                'data': norm_data(r[0]),
                '_dt': r[0],
                'esporte': norm_esporte(competicao),
                'tipster': TIPSTER,
                'casa': norm_casa(r[10]),
                'parceiro': PARCEIRO,
                'aposta': norm_categoria(competicao, mercado),
                'descricao': norm_descricao(competicao, mercado, r[3], r[4]),
                'stake': fmt_stake(r[6]),
                'odd': norm_odd(r[7]),
                'resultado': norm_resultado(r[8]),
                '_lucro': _para_float(r[9]),        # só para a conferência do DRY
                '_competicao': limpa(competicao),
                '_mercado': limpa(mercado),
            })
    return numerar(out)


def numerar(rows: list[dict]) -> list[dict]:
    """Código RC<aaaamm>-<n>. Numeração MENSAL (reinicia em 1 a cada mês), na
    ordem das linhas da planilha dentro do mês."""
    contador: dict[str, int] = defaultdict(int)
    for r in rows:
        mes = f"{r['_dt']:%Y%m}"
        contador[mes] += 1
        r['codigo'] = f'{PREFIXO}{mes}-{contador[mes]}'
        r['_mes'] = mes
    return rows


# ---------- assinatura (idêntica a repository._assinatura) ----------
# COM código o hash é `ID|casa|parceiro|codigo` — o conteúdo NÃO entra. É o que
# faz o bot casar a linha certa quando reprocessar o mesmo número.
def assinatura(r: dict) -> str:
    raw = '|'.join(['ID', r['casa'], r['parceiro'], r['codigo']])
    return hashlib.sha256(raw.encode()).hexdigest()[:20]


def estado_extracao(resultado: str, odd: str) -> str:
    """Espelha repository.estado_extracao: desde a s259 a odd só é exigida onde o
    P/L depende dela (W/HW); L/V/HL nascem `resolvida`."""
    if resultado not in VALID:
        return 'aberta'
    if resultado in ('W', 'HW'):
        return 'resolvida' if (_para_float(odd) or 0) > 0 else 'aberta'
    return 'resolvida'


def carregar_env():
    for line in open(ENV_PATH, encoding='utf-8'):
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


async def importar(rows: list[dict]):
    import asyncpg
    url = os.environ['DATABASE_URL'].replace('postgres://', 'postgresql://', 1)
    casas = sorted({r['casa'] for r in rows})

    registros = [(
        DONO, r['casa'], r['parceiro'], assinatura(r), r['codigo'],
        r['data'], r['esporte'], r['tipster'], r['aposta'], r['descricao'],
        r['stake'], r['odd'], r['resultado'] or None,
        estado_extracao(r['resultado'], r['odd']),
        None, None, ORIGEM,                       # confianca, stake_usd, origem
    ) for r in rows]

    last_err = None
    for tentativa in range(1, 4):
        try:
            conn = await asyncpg.connect(url, command_timeout=120)
            try:
                async with conn.transaction():
                    # idempotente: reimportar não acumula (limpa só o que ESTE
                    # import escreveu — captura do bot/extensão tem outra origem)
                    apagadas = await conn.execute(
                        'DELETE FROM bilhetes WHERE dono=$1 AND origem=$2', DONO, ORIGEM)
                    print(f'  [tentativa {tentativa}] limpou import anterior: {apagadas}')
                    await conn.executemany(
                        """
                        INSERT INTO bilhetes
                            (dono, casa, parceiro, assinatura, codigo_bilhete, data, esporte,
                             tipster, aposta, descricao, stake, odd, resultado,
                             extraction_state, confianca, stake_usd, origem)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17)
                        ON CONFLICT (dono, casa, parceiro, assinatura) DO NOTHING
                        """,
                        registros,
                    )
                    # uma conta `Padrão` POR CASA: 5 linhas no Painel de Contas,
                    # cada uma com custo próprio
                    for casa in casas:
                        await conn.execute(
                            """INSERT INTO parceiros (dono, casa, nome) VALUES ($1,$2,$3)
                               ON CONFLICT (dono, casa, nome) DO NOTHING""",
                            DONO, casa, PARCEIRO)
                    # feed ordena por criado_em DESC; num import não existe
                    # "envio" → ancora na data da aposta para sair cronológico
                    await conn.execute(
                        """
                        WITH ordered AS (
                            SELECT id,
                                   ROW_NUMBER() OVER (ORDER BY to_date(data,'DD/MM/YYYY') ASC,
                                                               id ASC) AS rn,
                                   COUNT(*) OVER () AS total
                            FROM bilhetes WHERE dono=$1 AND origem=$2
                        )
                        UPDATE bilhetes b
                        SET criado_em = NOW() - ((o.total - o.rn) * INTERVAL '1 second')
                        FROM ordered o WHERE b.id = o.id
                        """, DONO, ORIGEM)
                n = await conn.fetchval('SELECT COUNT(*) FROM bilhetes WHERE dono=$1', DONO)
                nc = await conn.fetchval(
                    'SELECT COUNT(DISTINCT casa) FROM parceiros WHERE dono=$1', DONO)
                np = await conn.fetchval('SELECT COUNT(*) FROM parceiros WHERE dono=$1', DONO)
                print(f'\nOK — bilhetes dono={DONO}={n} | casas={nc} | contas={np}')
                return
            finally:
                await conn.close()
        except Exception as e:                       # noqa: proxy instável → retry
            last_err = e
            print(f'  [tentativa {tentativa}] falhou: {type(e).__name__}: {e}')
    raise SystemExit(f'import falhou após 3 tentativas: {last_err}')


# ---------- relatório do DRY ----------
def _pl_derivado(stake: str, odd: str, res: str):
    s = _para_float(stake) or 0.0
    o = _para_float(odd)
    if res == 'L':
        return -s
    if res == 'V':
        return 0.0
    if res == 'HL':
        return -s / 2
    if o is None:
        return None
    if res == 'W':
        return s * (o - 1)
    if res == 'HW':
        return s * (o - 1) / 2
    return None


def _relatorio(rows: list[dict]):
    print(f'DONO={DONO} | tipster={TIPSTER!r} | conta={PARCEIRO!r} por casa | '
          f'linhas: {len(rows)}')

    print('\n— por aba —')
    for aba in ABAS:
        sub = [r for r in rows if r['aba'] == aba]
        if not sub:
            continue
        datas = sorted(r['_dt'] for r in sub)
        print(f'  {aba:<13} {len(sub):>4} | {datas[0]:%d/%m/%Y} → {datas[-1]:%d/%m/%Y} '
              f'| códigos {sub[0]["codigo"]} … {sub[-1]["codigo"]}')

    for campo in ('casa', 'esporte', 'aposta'):
        print(f'\n{campo}:', dict(Counter(r[campo] for r in rows).most_common()))
    print('\nresultado:',
          dict(Counter(r['resultado'] or '(aberta)' for r in rows).most_common()))

    nc = Counter(r['esporte'] for r in rows if r['esporte'] in NAO_CANONICOS)
    if nc:
        print(f'\n⚠ esporte(s) FORA do MASTER_ESPORTES §7 (entram verbatim; a rota '
              f'/taxonomia serve a união com a base do dono): {dict(nc)}')

    sigs = [assinatura(r) for r in rows]
    cods = [r['codigo'] for r in rows]
    print(f'\ncódigos: {len(set(cods))} únicos de {len(cods)}')
    print(f'assinaturas: {len(set(sigs))} únicas de {len(sigs)}')

    print('extraction_state:', dict(Counter(
        estado_extracao(r['resultado'], r['odd']) for r in rows)))

    sem_odd = [r for r in rows if not r['odd']]
    if sem_odd:
        print(f'\n⚠ {len(sem_odd)} linha(s) sem odd')
    sem_stake = [r for r in rows if (_para_float(r['stake']) or 0) <= 0]
    if sem_stake:
        print(f'\n⚠ {len(sem_stake)} linha(s) com stake 0 ou vazio: gravadas, porém '
              f'INVISÍVEIS no dashboard (dashboard_rows corta stake <= 0):')
        for r in sem_stake[:10]:
            print(f'    {r["codigo"]} | {r["data"]} | {r["descricao"][:50]} | '
                  f'stake={r["stake"]!r}')

    # conferência contra a coluna Lucro da própria planilha, linha a linha: se o
    # P/L derivado divergir, a classificação ou a normalização quebrou.
    liq = [r for r in rows if r['resultado']]
    div = [r for r in liq
           if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is None
           or abs(v - (r['_lucro'] or 0)) > 0.02]
    print(f'\nP/L derivado × coluna "Lucro Obtido (Un)" da planilha: '
          f'{len(div)} divergência(s) em {len(liq)} liquidadas')
    for r in div[:12]:
        print(f'    {r["codigo"]} | {r["data"]} | {r["resultado"]} | '
              f'{r["descricao"][:40]} | u={r["stake"]} @{r["odd"]} | '
              f'planilha={r["_lucro"]} | derivado='
              f'{_pl_derivado(r["stake"], r["odd"], r["resultado"])}')

    turnover = sum(_para_float(r['stake']) or 0 for r in rows)
    to_liq = sum(_para_float(r['stake']) or 0 for r in liq)
    pl = sum(v for r in liq
             if (v := _pl_derivado(r['stake'], r['odd'], r['resultado'])) is not None)
    print(f'\nturnover total (u):      {turnover:>10,.2f}')
    print(f'turnover liquidado (u):  {to_liq:>10,.2f}')
    print(f'P/L derivado (u):        {pl:>10,.2f}   '
          f'(ROI {pl / to_liq * 100:.1f}% sobre o liquidado)')

    print('\n— pares (Competição, Mercado) → (esporte, categoria) —')
    pares = Counter((r['_competicao'], r['_mercado'], r['esporte'], r['aposta'])
                    for r in rows)
    for (comp, merc, esp, cat), n in sorted(pares.items(), key=lambda kv: -kv[1]):
        alerta = '  ⚠ Outros' if cat == 'Outros' else ''
        print(f'  {comp:<15} + {merc:<11} → {esp:<14} / {cat:<15} ({n}){alerta}')

    print('\n=== 16 AMOSTRAS (Código|Data|Esporte|Casa|Aposta|Descrição|Stake|Odd|Res) ===')
    step = max(1, len(rows) // 16)
    for r in rows[::step][:16]:
        print(' | '.join([r['codigo'], r['data'], r['esporte'], r['casa'], r['aposta'],
                          r['descricao'][:42], r['stake'], r['odd'],
                          r['resultado'] or '—']))

    ultimo = max((r for r in rows if r['_mes'] == '202608'),
                 key=lambda r: int(r['codigo'].rsplit('-', 1)[1]), default=None)
    if ultimo:
        n = int(ultimo['codigo'].rsplit('-', 1)[1])
        print(f'\n>>> CONTADOR DO BOT — data/<tipster>/contador.json: '
              f'{{"2026-08": {n}}}  (próxima do grupo = #{n + 1})')


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--go', action='store_true')
    args = ap.parse_args()
    rows = carregar_rows(args.xlsx)
    _relatorio(rows)
    if not args.go:
        print('\n[DRY] nada escrito. Rode com --go para importar.')
        return
    carregar_env()
    asyncio.run(importar(rows))


if __name__ == '__main__':
    main()
